from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from types import SimpleNamespace

import colour
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from colour import SpectralDistribution, SpectralShape
from colour.quality import colour_fidelity_index
from pysilsub.CIE import get_CIE_1924_photopic_vl, get_CIES026_action_spectra
from scipy.optimize import minimize
from scipy.stats import friedmanchisquare, wilcoxon


ROOT = Path(r"D:\codex\tmp\led_paper")
OUT = ROOT / "analysis_output"
FIG = OUT / "figures"
DATA = OUT / "source_data"
WORKBOOK = Path(r"C:\Users\13956\Downloads\附录.xlsx")

WAVELENGTHS = np.arange(380.0, 781.0, 1.0)
CHANNELS = ["Blue", "Green", "Red", "Warm White", "Cold White"]
CHANNELS_ZH = ["蓝光", "绿光", "红光", "暖白光", "冷白光"]
CHANNEL_COLORS = ["#3976AF", "#3E9A75", "#C9584E", "#E2A84B", "#6D83B5"]
COND = ["A 优化光照", "B 普通LED", "C 黑暗"]
COND_SHORT = ["A", "B", "C"]
COND_COLORS = ["#3976AF", "#8A8F98", "#D28D3D"]

mpl.rcParams.update(
    {
        "font.family": "sans-serif",
        "font.sans-serif": ["Microsoft YaHei", "Arial", "DejaVu Sans", "sans-serif"],
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
        "font.size": 7.2,
        "axes.labelsize": 7.5,
        "axes.titlesize": 8.2,
        "xtick.labelsize": 6.5,
        "ytick.labelsize": 6.5,
        "legend.fontsize": 6.4,
        "axes.spines.right": False,
        "axes.spines.top": False,
        "axes.linewidth": 0.7,
        "legend.frameon": False,
        "lines.linewidth": 1.35,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "savefig.facecolor": "white",
    }
)


def parse_wavelength(value) -> float:
    match = re.match(r"\s*([0-9]+(?:\.[0-9]+)?)", str(value))
    if not match:
        raise ValueError(f"Cannot parse wavelength: {value!r}")
    return float(match.group(1))


def load_spectral_sheet(ws, columns: list[int]) -> tuple[np.ndarray, np.ndarray]:
    rows = list(ws.iter_rows(min_row=2, max_row=402, max_col=max(columns), values_only=True))
    wl = np.array([parse_wavelength(row[0]) for row in rows])
    values = np.array([[float(row[c - 1] or 0.0) for c in columns] for row in rows])
    if not np.array_equal(wl, WAVELENGTHS):
        raise ValueError("Expected a 380-780 nm grid at 1 nm intervals.")
    return wl, values


@dataclass(frozen=True)
class SpectralMetrics:
    cct: float
    duv: float
    rf: float
    rg: float
    mel_der: float


class MetricEngine:
    def __init__(self):
        action = get_CIES026_action_spectra(binwidth=1)
        self.melanopic = action["mel"].to_numpy(dtype=float)
        self.photopic = get_CIE_1924_photopic_vl(binwidth=1).iloc[:, 0].to_numpy(dtype=float)
        d65 = colour.SDS_ILLUMINANTS["D65"].copy().align(SpectralShape(380, 780, 1))
        self.d65 = d65.values.astype(float)
        self.d65_ratio = self._alpha_photopic_ratio(self.d65)
        cmfs = colour.MSDS_CMFS["CIE 1931 2 Degree Standard Observer"].copy().align(
            SpectralShape(380, 780, 1)
        )
        self.cmfs = cmfs.values.astype(float)

    def _alpha_photopic_ratio(self, spd: np.ndarray) -> float:
        den = np.trapezoid(spd * self.photopic, WAVELENGTHS)
        if den <= 0:
            return float("nan")
        return float(np.trapezoid(spd * self.melanopic, WAVELENGTHS) / den)

    def mel_der(self, spd: np.ndarray) -> float:
        return self._alpha_photopic_ratio(spd) / self.d65_ratio

    def xyz(self, spd: np.ndarray) -> np.ndarray:
        return np.trapezoid(spd[:, None] * self.cmfs, WAVELENGTHS, axis=0)

    def cct_duv_fast(self, spd: np.ndarray) -> tuple[float, float]:
        xyz = self.xyz(spd)
        xy = colour.XYZ_to_xy(xyz)
        uv = colour.xy_to_UCS_uv(xy)
        cct, duv = colour.temperature.uv_to_CCT_Robertson1968(uv)
        return float(cct), float(duv)

    def tm30(self, spd: np.ndarray):
        sd = SpectralDistribution(dict(zip(WAVELENGTHS, np.maximum(spd, 0.0))))
        return colour_fidelity_index(
            sd, additional_data=True, method="ANSI/IES TM-30-18"
        )

    def all_metrics(self, spd: np.ndarray) -> tuple[SpectralMetrics, object]:
        spec = self.tm30(spd)
        metrics = SpectralMetrics(
            cct=float(spec.CCT),
            duv=float(spec.D_uv),
            rf=float(spec.R_f),
            rg=float(spec.R_g),
            mel_der=float(self.mel_der(spd)),
        )
        return metrics, spec


def normalize_spd(spd: np.ndarray) -> np.ndarray:
    area = np.trapezoid(spd, WAVELENGTHS)
    if area <= 0:
        raise ValueError("SPD integral must be positive.")
    return spd / area


def spectral_relative_error(test: np.ndarray, target: np.ndarray) -> float:
    a = normalize_spd(test)
    b = normalize_spd(target)
    return float(np.linalg.norm(a - b) / np.linalg.norm(b))


def save_figure(fig: mpl.figure.Figure, stem: str) -> None:
    FIG.mkdir(parents=True, exist_ok=True)
    fig.savefig(FIG / f"{stem}.svg", bbox_inches="tight")
    fig.savefig(FIG / f"{stem}.pdf", bbox_inches="tight")
    fig.savefig(FIG / f"{stem}.png", dpi=600, bbox_inches="tight")
    fig.savefig(FIG / f"{stem}.tiff", dpi=600, bbox_inches="tight")
    plt.close(fig)


def panel_label(ax, label: str, x: float = -0.12, y: float = 1.06) -> None:
    ax.text(
        x,
        y,
        label,
        transform=ax.transAxes,
        fontsize=8.5,
        fontweight="bold",
        ha="left",
        va="top",
    )


def style_axis(ax) -> None:
    ax.tick_params(direction="out", length=2.4, width=0.65)
    ax.grid(False)


def extract_gamut(spec) -> tuple[np.ndarray, np.ndarray]:
    test = np.asarray(spec.averages_test)
    ref = np.asarray(spec.averages_reference)
    if test.ndim == 2 and test.shape[1] == 2:
        return test, ref
    if test.ndim == 2 and test.shape[1] >= 3:
        return test[:, -2:], ref[:, -2:]
    raise ValueError(f"Unexpected TM-30 gamut data shape: {test.shape}")


def plot_gamut(ax, spec, title: str) -> None:
    test, ref = extract_gamut(spec)
    test = np.vstack([test, test[0]])
    ref = np.vstack([ref, ref[0]])
    ax.plot(ref[:, 0], ref[:, 1], color="#9AA0A6", ls="--", lw=1.0, label="参考光源")
    ax.plot(test[:, 0], test[:, 1], color="#3976AF", lw=1.4, label="测试光源")
    ax.fill(test[:, 0], test[:, 1], color="#3976AF", alpha=0.12)
    ax.axhline(0, color="#D6D8DB", lw=0.5)
    ax.axvline(0, color="#D6D8DB", lw=0.5)
    ax.set_aspect("equal", adjustable="datalim")
    ax.set_xlabel("CAM02-UCS a′")
    ax.set_ylabel("CAM02-UCS b′")
    ax.set_title(title, loc="left", fontweight="bold")
    style_axis(ax)


def optimise_modes(engine: MetricEngine, channel_spds: np.ndarray):
    cache: dict[tuple[float, ...], tuple[SpectralMetrics, object]] = {}

    def evaluate(w: np.ndarray):
        w = np.asarray(w, dtype=float)
        key = tuple(np.round(w, 11))
        if key not in cache:
            cache[key] = engine.all_metrics(channel_spds @ w)
        return cache[key]

    bounds = [(0.0, 1.0)] * 5
    eq = {"type": "eq", "fun": lambda w: np.sum(w) - 1.0}

    day_constraints = [
        eq,
        {"type": "ineq", "fun": lambda w: evaluate(w)[0].cct - 5500.0},
        {"type": "ineq", "fun": lambda w: 6500.0 - evaluate(w)[0].cct},
        {"type": "ineq", "fun": lambda w: evaluate(w)[0].rg - 95.0},
        {"type": "ineq", "fun": lambda w: 105.0 - evaluate(w)[0].rg},
        {"type": "ineq", "fun": lambda w: evaluate(w)[0].rf - 88.0},
    ]
    night_constraints = [
        eq,
        {"type": "ineq", "fun": lambda w: evaluate(w)[0].cct - 2500.0},
        {"type": "ineq", "fun": lambda w: 3500.0 - evaluate(w)[0].cct},
        {"type": "ineq", "fun": lambda w: evaluate(w)[0].rf - 80.0},
    ]

    rng = np.random.default_rng(20250822)
    day_seeds = [
        np.array([0.3517, 0.1400, 0.0, 0.1107, 0.3976]),
        np.array([0.15, 0.10, 0.05, 0.20, 0.50]),
    ]
    night_seeds = [
        np.array([0.1923, 0.1753, 0.1087, 0.5009, 0.0229]),
        np.array([0.03, 0.10, 0.20, 0.64, 0.03]),
    ]
    candidates = rng.dirichlet(np.ones(5), size=320)
    fast = []
    for w in candidates:
        cct, _ = engine.cct_duv_fast(channel_spds @ w)
        fast.append((w, cct, engine.mel_der(channel_spds @ w)))
    day_pool = sorted((x for x in fast if 5400 <= x[1] <= 6600), key=lambda x: abs(x[1] - 6200))
    night_pool = sorted((x for x in fast if 2400 <= x[1] <= 3600), key=lambda x: x[2])
    day_seeds.extend([x[0] for x in day_pool[:1]])
    night_seeds.extend([x[0] for x in night_pool[:1]])

    day_results = []
    for seed_index, x0 in enumerate(day_seeds):
        print(f"Day-mode TM-30 refinement {seed_index + 1}/{len(day_seeds)}", flush=True)
        res = minimize(
            lambda w: -evaluate(w)[0].rf,
            x0,
            method="SLSQP",
            bounds=bounds,
            constraints=day_constraints,
            options={"ftol": 2e-6, "maxiter": 45, "eps": 2e-5, "disp": False},
        )
        m, spec = evaluate(res.x)
        feasible = (
            abs(np.sum(res.x) - 1) < 2e-5
            and np.min(res.x) >= -2e-6
            and 5499.5 <= m.cct <= 6500.5
            and 94.98 <= m.rg <= 105.02
            and m.rf >= 87.98
        )
        if feasible:
            day_results.append((res, m, spec))
    if not day_results:
        raise RuntimeError("No feasible day-mode solution found.")
    day = max(day_results, key=lambda x: x[1].rf)

    night_results = []
    for seed_index, x0 in enumerate(night_seeds):
        print(f"Night-mode TM-30 refinement {seed_index + 1}/{len(night_seeds)}", flush=True)
        res = minimize(
            lambda w: evaluate(w)[0].mel_der,
            x0,
            method="SLSQP",
            bounds=bounds,
            constraints=night_constraints,
            options={"ftol": 2e-7, "maxiter": 55, "eps": 2e-5, "disp": False},
        )
        m, spec = evaluate(res.x)
        feasible = (
            abs(np.sum(res.x) - 1) < 2e-5
            and np.min(res.x) >= -2e-6
            and 2499.5 <= m.cct <= 3500.5
            and m.rf >= 79.98
        )
        if feasible:
            night_results.append((res, m, spec))
    if not night_results:
        raise RuntimeError("No feasible night-mode solution found.")
    night = min(night_results, key=lambda x: x[1].mel_der)
    return day, night, cache


def optimise_daylight(engine: MetricEngine, channel_spds: np.ndarray, targets: np.ndarray):
    n = targets.shape[1]
    target_norm = np.column_stack([normalize_spd(targets[:, j]) for j in range(n)])
    target_metrics = []
    for j in range(n):
        cct, duv = engine.cct_duv_fast(targets[:, j])
        target_metrics.append((cct, duv, engine.mel_der(targets[:, j])))
    target_metrics = np.array(target_metrics)

    def local_loss(w, j, neighbours=()):
        mix = channel_spds @ w
        mix_norm = normalize_spd(mix)
        shape = np.linalg.norm(mix_norm - target_norm[:, j]) / np.linalg.norm(target_norm[:, j])
        cct, duv = engine.cct_duv_fast(mix)
        mel = engine.mel_der(mix)
        dcct = (cct - target_metrics[j, 0]) / 1000.0
        dduv = (duv - target_metrics[j, 1]) / 0.010
        dmel = (mel - target_metrics[j, 2]) / 0.20
        smooth = sum(float(np.sum((w - neighbour) ** 2)) for neighbour in neighbours)
        return 1.0 * shape**2 + 0.55 * dcct**2 + 0.18 * dduv**2 + 0.45 * dmel**2 + 0.12 * smooth

    weights = np.tile(np.array([0.08, 0.18, 0.03, 0.34, 0.37]), (n, 1))
    eq = {"type": "eq", "fun": lambda w: np.sum(w) - 1.0}
    bounds = [(0.0, 1.0)] * 5
    success = True
    messages = []
    for sweep in range(1):
        order = range(n)
        for j in order:
            neighbours = []
            if j > 0:
                neighbours.append(weights[j - 1].copy())
            if j < n - 1:
                neighbours.append(weights[j + 1].copy())
            res = minimize(
                lambda w, jj=j, nn=tuple(neighbours): local_loss(w, jj, nn),
                weights[j],
                method="SLSQP",
                bounds=bounds,
                constraints=[eq],
                options={"ftol": 2e-6, "maxiter": 20, "eps": 1e-5, "disp": False},
            )
            weights[j] = np.clip(res.x, 0, 1)
            weights[j] /= weights[j].sum()
            success = success and bool(res.success)
            if not res.success:
                messages.append(f"{j}:{res.message}")
        print(f"Daylight smoothing sweep {sweep + 1}/1", flush=True)

    objective_value = float(
        np.mean([local_loss(weights[j], j, ()) for j in range(n)])
        + 0.12 * np.sum(np.diff(weights, axis=0) ** 2) / max(n - 1, 1)
    )
    result = SimpleNamespace(
        success=success,
        message="; ".join(messages) if messages else "Sequential SLSQP converged",
        fun=objective_value,
        x=weights.ravel(),
    )
    if not success:
        raise RuntimeError(f"Daylight optimization failed: {result.message}")
    rows = []
    mixes = []
    for j in range(n):
        mix = channel_spds @ weights[j]
        mixes.append(mix)
        cct, duv = engine.cct_duv_fast(mix)
        rows.append(
            {
                "target_cct": target_metrics[j, 0],
                "synth_cct": cct,
                "target_duv": target_metrics[j, 1],
                "synth_duv": duv,
                "target_mel_der": target_metrics[j, 2],
                "synth_mel_der": engine.mel_der(mix),
                "relative_spectral_error": spectral_relative_error(mix, targets[:, j]),
                "weight_step_l2": 0.0 if j == 0 else float(np.linalg.norm(weights[j] - weights[j - 1])),
            }
        )
    return result, weights, np.column_stack(mixes), pd.DataFrame(rows)


def sleep_metrics(values: np.ndarray) -> dict[str, float]:
    stages = values[np.isfinite(values)].astype(int)
    if stages.size == 0:
        raise ValueError("Empty sleep record")
    sleep = stages != 4
    first_sleep = int(np.flatnonzero(sleep)[0])
    tst = 0.5 * sleep.sum()
    tib = 0.5 * len(stages)
    sol = 0.5 * first_sleep
    n3 = 100.0 * np.sum(stages == 3) / sleep.sum()
    rem = 100.0 * np.sum(stages == 5) / sleep.sum()
    awakenings = int(np.sum((stages[1:] == 4) & (stages[:-1] != 4) & (np.arange(1, len(stages)) > first_sleep)))
    return {
        "TST": tst,
        "SE": 100.0 * tst / tib,
        "SOL": sol,
        "N3_pct": n3,
        "REM_pct": rem,
        "Awakenings": float(awakenings),
        "TIB": tib,
        "epochs": float(len(stages)),
    }


def bootstrap_mean_difference(a: np.ndarray, b: np.ndarray, seed: int) -> tuple[float, float, float]:
    d = np.asarray(a) - np.asarray(b)
    rng = np.random.default_rng(seed)
    indices = rng.integers(0, len(d), size=(30000, len(d)))
    boot = d[indices].mean(axis=1)
    lo, hi = np.quantile(boot, [0.025, 0.975])
    return float(d.mean()), float(lo), float(hi)


def rank_biserial_from_differences(a: np.ndarray, b: np.ndarray) -> float:
    d = np.asarray(a) - np.asarray(b)
    d = d[d != 0]
    if len(d) == 0:
        return 0.0
    ranks = pd.Series(np.abs(d)).rank(method="average").to_numpy()
    pos = ranks[d > 0].sum()
    neg = ranks[d < 0].sum()
    return float((pos - neg) / (pos + neg))


def holm_adjust(pvalues: list[float]) -> list[float]:
    p = np.asarray(pvalues, dtype=float)
    order = np.argsort(p)
    adjusted_sorted = np.maximum.accumulate((len(p) - np.arange(len(p))) * p[order])
    adjusted_sorted = np.clip(adjusted_sorted, 0, 1)
    adjusted = np.empty_like(p)
    adjusted[order] = adjusted_sorted
    return adjusted.tolist()


def analyse_sleep(ws):
    raw_rows = list(ws.iter_rows(min_row=3, max_col=33, values_only=True))
    raw = np.array(
        [[np.nan if value is None else float(value) for value in row] for row in raw_rows],
        dtype=float,
    )
    records = []
    for subject in range(11):
        for night in range(3):
            col = subject * 3 + night
            vals = raw[:, col]
            metrics = sleep_metrics(vals)
            records.append({"subject": subject + 1, "condition": COND_SHORT[night], **metrics})
    frame = pd.DataFrame(records)
    endpoints = ["TST", "SE", "SOL", "N3_pct", "REM_pct", "Awakenings"]
    omnibus = []
    pairs = [(0, 1), (0, 2), (1, 2)]
    pair_rows = []
    for endpoint in endpoints:
        matrix = frame.pivot(index="subject", columns="condition", values=endpoint)[COND_SHORT].to_numpy()
        stat, p = friedmanchisquare(matrix[:, 0], matrix[:, 1], matrix[:, 2])
        omnibus.append(
            {
                "endpoint": endpoint,
                "chi2": float(stat),
                "p": float(p),
                "kendalls_w": float(stat / (11 * 2)),
            }
        )
        if p < 0.05:
            raw = []
            local = []
            for i, j in pairs:
                wr = wilcoxon(matrix[:, i], matrix[:, j], alternative="two-sided", method="auto")
                mean_diff, lo, hi = bootstrap_mean_difference(matrix[:, i], matrix[:, j], 8000 + i * 10 + j)
                raw.append(float(wr.pvalue))
                local.append(
                    {
                        "endpoint": endpoint,
                        "comparison": f"{COND_SHORT[i]}-{COND_SHORT[j]}",
                        "W": float(wr.statistic),
                        "p_raw": float(wr.pvalue),
                        "mean_difference": mean_diff,
                        "ci95_low": lo,
                        "ci95_high": hi,
                        "rank_biserial": rank_biserial_from_differences(matrix[:, i], matrix[:, j]),
                    }
                )
            adjusted = holm_adjust(raw)
            for row, adj in zip(local, adjusted):
                row["p_holm"] = adj
                pair_rows.append(row)
    desc = (
        frame.groupby("condition")[endpoints]
        .agg(["mean", "std", "median"])
        .reset_index()
    )
    return frame, desc, pd.DataFrame(omnibus), pd.DataFrame(pair_rows)


def figure_problem1(spd, metrics, spec, engine):
    fig = plt.figure(figsize=(6.65, 3.65))
    gs = fig.add_gridspec(1, 3, width_ratios=[1.55, 1.0, 0.9], wspace=0.42)
    ax = fig.add_subplot(gs[0, 0])
    ax.fill_between(WAVELENGTHS, spd, color="#3976AF", alpha=0.16)
    ax.plot(WAVELENGTHS, spd, color="#3976AF", lw=1.35)
    ax2 = ax.twinx()
    ax2.plot(WAVELENGTHS, engine.photopic / engine.photopic.max(), color="#6F747C", lw=0.9, ls="--", label="明视觉 V(λ)")
    ax2.plot(WAVELENGTHS, engine.melanopic / engine.melanopic.max(), color="#D28D3D", lw=1.0, label="黑视素 s_mel(λ)")
    ax.set_xlabel("波长 / nm")
    ax.set_ylabel("相对光谱功率")
    ax2.set_ylabel("")
    ax.set_title("实测光谱与视觉/节律加权", loc="left", fontweight="bold")
    lines, labels = ax2.get_legend_handles_labels()
    ax2.legend(lines, labels, loc="upper right")
    style_axis(ax)
    style_axis(ax2)
    panel_label(ax, "a")

    axg = fig.add_subplot(gs[0, 1])
    plot_gamut(axg, spec, "TM-30 16色相区色域")
    axg.legend(loc="lower left")
    panel_label(axg, "b")

    axm = fig.add_subplot(gs[0, 2])
    axm.axis("off")
    labels = ["CCT", "Duv", "Rf", "Rg", "mel-DER"]
    values = [f"{metrics.cct:.0f} K", f"{metrics.duv:+.5f}", f"{metrics.rf:.2f}", f"{metrics.rg:.2f}", f"{metrics.mel_der:.3f}"]
    y = np.linspace(0.75, 0.15, len(labels))
    for yy, lab, val in zip(y, labels, values):
        axm.text(0.02, yy, lab, color="#6F747C", ha="left", va="center", fontsize=7)
        axm.text(0.98, yy, val, color="#202124", ha="right", va="center", fontsize=9, fontweight="bold")
        axm.plot([0.02, 0.98], [yy - 0.07, yy - 0.07], color="#E1E3E6", lw=0.6)
    axm.set_title("标准化计算结果", loc="left", fontweight="bold")
    panel_label(axm, "c")
    save_figure(fig, "fig1_problem1_standard_metrics")


def figure_problem2(channel_spds, day, night):
    day_res, day_m, day_spec = day
    night_res, night_m, night_spec = night
    spectra = [channel_spds @ day_res.x, channel_spds @ night_res.x]
    fig = plt.figure(figsize=(6.65, 6.0))
    gs = fig.add_gridspec(2, 3, width_ratios=[1.55, 1.0, 1.0], hspace=0.42, wspace=0.40)

    ax0 = fig.add_subplot(gs[:, 0])
    for i in range(5):
        ax0.plot(WAVELENGTHS, normalize_spd(channel_spds[:, i]), color=CHANNEL_COLORS[i], label=CHANNELS_ZH[i], lw=1.15)
    ax0.set_xlabel("波长 / nm")
    ax0.set_ylabel("面积归一化光谱功率")
    ax0.set_title("五通道光谱基底", loc="left", fontweight="bold")
    ax0.legend(loc="upper right", ncol=1)
    style_axis(ax0)
    panel_label(ax0, "a")

    for row, (name, w, m, spec, spd, accent) in enumerate(
        [
            ("日间模式", day_res.x, day_m, day_spec, spectra[0], "#3976AF"),
            ("夜间模式", night_res.x, night_m, night_spec, spectra[1], "#D28D3D"),
        ]
    ):
        ax = fig.add_subplot(gs[row, 1])
        bars = ax.bar(np.arange(5), w, color=CHANNEL_COLORS, width=0.72)
        ax.set_ylim(0, max(0.62, float(np.max(w)) * 1.18))
        ax.set_xticks(np.arange(5), ["B", "G", "R", "WW", "CW"])
        ax.set_ylabel("通道权重")
        ax.set_title(name + "权重", loc="left", fontweight="bold")
        for bar, val in zip(bars, w):
            if val >= 0.015:
                ax.text(bar.get_x() + bar.get_width() / 2, val + 0.015, f"{val:.3f}", ha="center", va="bottom", fontsize=5.8)
        style_axis(ax)
        panel_label(ax, "b" if row == 0 else "d")

        axg = fig.add_subplot(gs[row, 2])
        plot_gamut(axg, spec, f"{name}色域")
        axg.text(
            0.02,
            0.02,
            f"CCT {m.cct:.0f} K\nRf {m.rf:.2f}  Rg {m.rg:.2f}\nmel-DER {m.mel_der:.3f}",
            transform=axg.transAxes,
            ha="left",
            va="bottom",
            fontsize=6.1,
            color="#202124",
        )
        panel_label(axg, "c" if row == 0 else "e")
    save_figure(fig, "fig2_day_night_optimization")


def figure_problem3(times, weights, diagnostics):
    x = np.arange(len(times))
    fig = plt.figure(figsize=(6.65, 5.75))
    gs = fig.add_gridspec(3, 2, height_ratios=[1.2, 1.0, 0.9], hspace=0.48, wspace=0.40)
    axw = fig.add_subplot(gs[0, :])
    for i in range(5):
        axw.plot(x, weights[:, i], marker="o", ms=2.5, color=CHANNEL_COLORS[i], label=CHANNELS_ZH[i])
    axw.set_xticks(x, times, rotation=0, rotation_mode="anchor")
    axw.set_ylabel("通道权重")
    axw.set_title("全天五通道控制轨迹", loc="left", fontweight="bold")
    axw.legend(loc="upper center", bbox_to_anchor=(0.5, 1.23), ncol=5)
    style_axis(axw)
    panel_label(axw, "a")

    axc = fig.add_subplot(gs[1, 0])
    axc.plot(x, diagnostics["target_cct"], color="#8A8F98", ls="--", marker="o", ms=2.5, label="太阳目标")
    axc.plot(x, diagnostics["synth_cct"], color="#3976AF", marker="o", ms=2.5, label="LED合成")
    axc.set_xticks(x[::2], np.array(times)[::2])
    axc.set_ylabel("CCT / K")
    axc.set_title("相关色温跟踪", loc="left", fontweight="bold")
    axc.legend(loc="best")
    style_axis(axc)
    panel_label(axc, "b")

    axm = fig.add_subplot(gs[1, 1])
    axm.plot(x, diagnostics["target_mel_der"], color="#8A8F98", ls="--", marker="o", ms=2.5, label="太阳目标")
    axm.plot(x, diagnostics["synth_mel_der"], color="#D28D3D", marker="o", ms=2.5, label="LED合成")
    axm.set_xticks(x[::2], np.array(times)[::2])
    axm.set_ylabel("mel-DER")
    axm.set_title("节律效应跟踪", loc="left", fontweight="bold")
    axm.legend(loc="best")
    style_axis(axm)
    panel_label(axm, "c")

    axd = fig.add_subplot(gs[2, 0])
    axd.plot(x, diagnostics["synth_duv"] - diagnostics["target_duv"], color="#3E9A75", marker="o", ms=2.5)
    axd.axhline(0, color="#9AA0A6", lw=0.8, ls="--")
    axd.set_xticks(x[::2], np.array(times)[::2])
    axd.set_ylabel("ΔDuv")
    axd.set_title("色偏差残差", loc="left", fontweight="bold")
    style_axis(axd)
    panel_label(axd, "d")

    axe = fig.add_subplot(gs[2, 1])
    axe.plot(x, diagnostics["relative_spectral_error"], color="#C9584E", marker="o", ms=2.5, label="相对光谱误差")
    axe2 = axe.twinx()
    axe2.plot(x, diagnostics["weight_step_l2"], color="#6F747C", marker="s", ms=2.2, label="相邻权重步长")
    axe.set_xticks(x[::2], np.array(times)[::2])
    axe.set_ylabel("相对光谱误差")
    axe2.set_ylabel("权重步长 L2")
    axe.set_title("拟合误差与控制平滑性", loc="left", fontweight="bold")
    lines = axe.get_lines() + axe2.get_lines()
    axe.legend(lines, [line.get_label() for line in lines], loc="best")
    style_axis(axe)
    style_axis(axe2)
    panel_label(axe, "e")
    save_figure(fig, "fig3_full_day_control")


def figure_representative(times, targets, mixes, diagnostics):
    selected = [0, 4, len(times) - 1]
    labels = [times[i] for i in selected]
    fig = plt.figure(figsize=(6.65, 5.2))
    gs = fig.add_gridspec(2, 3, height_ratios=[1.5, 0.75], hspace=0.16, wspace=0.34)
    for col, idx in enumerate(selected):
        target = normalize_spd(targets[:, idx])
        mix = normalize_spd(mixes[:, idx])
        ax = fig.add_subplot(gs[0, col])
        ax.plot(WAVELENGTHS, target, color="#8A8F98", ls="--", label="太阳目标")
        ax.plot(WAVELENGTHS, mix, color="#3976AF", label="LED合成")
        ax.set_title(labels[col], fontweight="bold")
        if col == 0:
            ax.set_ylabel("面积归一化光谱功率")
            ax.legend(loc="upper right")
        ax.set_xticklabels([])
        ax.text(
            0.04,
            0.05,
            f"CCT误差 {diagnostics.iloc[idx].synth_cct - diagnostics.iloc[idx].target_cct:+.0f} K\nmel-DER误差 {diagnostics.iloc[idx].synth_mel_der - diagnostics.iloc[idx].target_mel_der:+.3f}",
            transform=ax.transAxes,
            ha="left",
            va="bottom",
            fontsize=5.8,
        )
        style_axis(ax)
        panel_label(ax, chr(ord("a") + col))

        axr = fig.add_subplot(gs[1, col])
        residual = mix - target
        axr.axhline(0, color="#9AA0A6", lw=0.7)
        axr.fill_between(WAVELENGTHS, 0, residual, where=residual >= 0, color="#D28D3D", alpha=0.55, interpolate=True)
        axr.fill_between(WAVELENGTHS, 0, residual, where=residual < 0, color="#3976AF", alpha=0.50, interpolate=True)
        axr.set_xlabel("波长 / nm")
        if col == 0:
            axr.set_ylabel("有符号残差")
        style_axis(axr)
    save_figure(fig, "fig4_representative_spectral_fits")


def figure_sleep(frame, omnibus, pairwise):
    endpoints = ["TST", "SE", "SOL", "N3_pct", "REM_pct", "Awakenings"]
    display = ["TST / min", "SE / %", "SOL / min", "N3 / %", "REM / %", "觉醒次数"]
    fig = plt.figure(figsize=(6.65, 6.1))
    gs = fig.add_gridspec(3, 3, height_ratios=[1.0, 1.0, 0.9], hspace=0.48, wspace=0.38)
    for k, (endpoint, title) in enumerate(zip(endpoints, display)):
        ax = fig.add_subplot(gs[k // 3, k % 3])
        matrix = frame.pivot(index="subject", columns="condition", values=endpoint)[COND_SHORT].to_numpy()
        for row in matrix:
            ax.plot(np.arange(3), row, color="#B6BAC0", lw=0.65, alpha=0.65, zorder=1)
            ax.scatter(np.arange(3), row, c=COND_COLORS, s=9, zorder=2, edgecolors="white", linewidths=0.25)
        means = matrix.mean(axis=0)
        sem = matrix.std(axis=0, ddof=1) / math.sqrt(matrix.shape[0])
        ax.errorbar(np.arange(3), means, yerr=sem, color="#202124", marker="D", ms=3.3, lw=1.2, capsize=2.2, zorder=3)
        p = float(omnibus.loc[omnibus.endpoint == endpoint, "p"].iloc[0])
        w = float(omnibus.loc[omnibus.endpoint == endpoint, "kendalls_w"].iloc[0])
        ax.set_title(f"{title}\nFriedman p={p:.3f}, W={w:.2f}", loc="left", fontweight="bold")
        ax.set_xticks(np.arange(3), COND_SHORT)
        style_axis(ax)
        panel_label(ax, chr(ord("a") + k), x=-0.16, y=1.14)

    axe = fig.add_subplot(gs[2, :])
    if pairwise.empty:
        axe.axis("off")
        axe.text(0.5, 0.5, "无总体显著指标，未进行事后检验", ha="center", va="center")
    else:
        rows = pairwise.iloc[::-1].reset_index(drop=True)
        y = np.arange(len(rows))
        errors = np.vstack([rows.mean_difference - rows.ci95_low, rows.ci95_high - rows.mean_difference])
        colors = ["#C9584E" if p < 0.05 else "#8A8F98" for p in rows.p_holm]
        for idx, (yy, color) in enumerate(zip(y, colors)):
            axe.errorbar(
                rows.mean_difference.iloc[idx],
                yy,
                xerr=[[errors[0, idx]], [errors[1, idx]]],
                fmt="o",
                color=color,
                ecolor=color,
                markersize=4.2,
                elinewidth=1.4,
                capsize=2.5,
                zorder=3,
            )
        axe.axvline(0, color="#9AA0A6", ls="--", lw=0.8)
        labels = [f"{r.endpoint}: {r.comparison}  (Holm p={r.p_holm:.3f})" for r in rows.itertuples()]
        axe.set_yticks(y, labels)
        axe.set_xlabel("前者减后者的配对均值差（bootstrap 95% CI）")
        axe.set_title("总体显著指标的事后配对效应", loc="left", fontweight="bold")
        style_axis(axe)
    panel_label(axe, "g", x=-0.10, y=1.14)
    save_figure(fig, "fig5_sleep_paired_inference")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    DATA.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    engine = MetricEngine()

    _, p1_values = load_spectral_sheet(wb["Problem 1"], [2])
    p1_spd = p1_values[:, 0]
    p1_metrics, p1_spec = engine.all_metrics(p1_spd)

    _, channel_spds = load_spectral_sheet(wb["Problem 2_LED_SPD"], [2, 3, 4, 5, 6])
    ws3 = wb["Problem 3 SUN_SPD"]
    header3 = next(ws3.iter_rows(min_row=1, max_row=1, max_col=16, values_only=True))
    times_all = [header3[c - 1].strftime("%H:%M") for c in range(2, 17)]
    use_cols = list(range(5, 17))
    times = [header3[c - 1].strftime("%H:%M") for c in use_cols]
    _, sun = load_spectral_sheet(ws3, use_cols)
    checkpoint = OUT / "optimization_checkpoint.npz"
    checkpoint_diag = OUT / "optimization_diagnostics.csv"
    if checkpoint.exists() and checkpoint_diag.exists():
        saved = np.load(checkpoint)
        day_w = saved["day_weights"]
        night_w = saved["night_weights"]
        weights = saved["daylight_weights"]
        mixes = saved["daylight_mixes"]
        day_m, day_spec = engine.all_metrics(channel_spds @ day_w)
        night_m, night_spec = engine.all_metrics(channel_spds @ night_w)
        day = (SimpleNamespace(x=day_w, success=True, message="Loaded verified checkpoint"), day_m, day_spec)
        night = (SimpleNamespace(x=night_w, success=True, message="Loaded verified checkpoint"), night_m, night_spec)
        diagnostics = pd.read_csv(checkpoint_diag)
        daylight_result = SimpleNamespace(success=True, message="Loaded verified checkpoint", fun=float(saved["daylight_objective"]))
        metric_cache = {}
        print("Loaded optimization checkpoint", flush=True)
    else:
        day, night, metric_cache = optimise_modes(engine, channel_spds)
        daylight_result, weights, mixes, diagnostics = optimise_daylight(engine, channel_spds, sun)
        diagnostics.insert(0, "time", times)
        for i, ch in enumerate(CHANNELS):
            diagnostics[f"w_{ch.replace(' ', '_')}"] = weights[:, i]
        np.savez(
            checkpoint,
            day_weights=day[0].x,
            night_weights=night[0].x,
            daylight_weights=weights,
            daylight_mixes=mixes,
            daylight_objective=float(daylight_result.fun),
        )
        diagnostics.to_csv(checkpoint_diag, index=False, encoding="utf-8-sig")

    diagnostics["normalized_spd_rmse"] = [
        float(np.sqrt(np.mean((normalize_spd(mixes[:, j]) - normalize_spd(sun[:, j])) ** 2)))
        for j in range(len(times))
    ]

    sleep_frame, sleep_desc, omnibus, pairwise = analyse_sleep(wb["Problem 4"])

    p1_df = pd.DataFrame({"wavelength_nm": WAVELENGTHS, "spd": p1_spd})
    channel_df = pd.DataFrame(channel_spds, columns=CHANNELS)
    channel_df.insert(0, "wavelength_nm", WAVELENGTHS)
    sun_df = pd.DataFrame(sun, columns=times)
    sun_df.insert(0, "wavelength_nm", WAVELENGTHS)
    mix_df = pd.DataFrame(mixes, columns=times)
    mix_df.insert(0, "wavelength_nm", WAVELENGTHS)
    p1_df.to_csv(DATA / "problem1_spd.csv", index=False, encoding="utf-8-sig")
    channel_df.to_csv(DATA / "problem2_channel_spds.csv", index=False, encoding="utf-8-sig")
    sun_df.to_csv(DATA / "problem3_target_spds.csv", index=False, encoding="utf-8-sig")
    mix_df.to_csv(DATA / "problem3_synthesized_spds.csv", index=False, encoding="utf-8-sig")
    diagnostics.to_csv(DATA / "problem3_diagnostics.csv", index=False, encoding="utf-8-sig")
    sleep_frame.to_csv(DATA / "problem4_subject_metrics.csv", index=False, encoding="utf-8-sig")
    sleep_desc.to_csv(DATA / "problem4_descriptive.csv", index=False, encoding="utf-8-sig")
    omnibus.to_csv(DATA / "problem4_friedman.csv", index=False, encoding="utf-8-sig")
    pairwise.to_csv(DATA / "problem4_pairwise.csv", index=False, encoding="utf-8-sig")

    day_res, day_m, _ = day
    night_res, night_m, _ = night
    results = {
        "workbook_time_columns": times_all,
        "analysis_time_columns": times,
        "problem1": p1_metrics.__dict__,
        "problem2": {
            "day": {"weights": day_res.x.tolist(), **day_m.__dict__, "optimizer_success": bool(day_res.success), "optimizer_message": str(day_res.message)},
            "night": {"weights": night_res.x.tolist(), **night_m.__dict__, "optimizer_success": bool(night_res.success), "optimizer_message": str(night_res.message)},
            "tm30_evaluations": len(metric_cache),
        },
        "problem3": {
            "optimizer_success": bool(daylight_result.success),
            "optimizer_message": str(daylight_result.message),
            "objective": float(daylight_result.fun),
            "mean_relative_spectral_error": float(diagnostics.relative_spectral_error.mean()),
            "mean_normalized_spd_rmse": float(diagnostics.normalized_spd_rmse.mean()),
            "max_normalized_spd_rmse": float(diagnostics.normalized_spd_rmse.max()),
            "max_abs_cct_error": float(np.max(np.abs(diagnostics.synth_cct - diagnostics.target_cct))),
            "mean_abs_cct_error": float(np.mean(np.abs(diagnostics.synth_cct - diagnostics.target_cct))),
            "max_abs_mel_der_error": float(np.max(np.abs(diagnostics.synth_mel_der - diagnostics.target_mel_der))),
            "mean_abs_mel_der_error": float(np.mean(np.abs(diagnostics.synth_mel_der - diagnostics.target_mel_der))),
            "max_weight_step_l2": float(diagnostics.weight_step_l2.max()),
        },
        "problem4": {
            "record_count": int(len(sleep_frame)),
            "subject_count": int(sleep_frame.subject.nunique()),
            "condition_mapping_assumption": {"Night 1": "A", "Night 2": "B", "Night 3": "C"},
        },
    }
    (OUT / "results.json").write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    figure_problem1(p1_spd, p1_metrics, p1_spec, engine)
    figure_problem2(channel_spds, day, night)
    figure_problem3(times, weights, diagnostics)
    figure_representative(times, sun, mixes, diagnostics)
    figure_sleep(sleep_frame, omnibus, pairwise)
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
