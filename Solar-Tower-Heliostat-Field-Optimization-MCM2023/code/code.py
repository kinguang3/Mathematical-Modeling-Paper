# -*- coding: utf-8 -*-
from __future__ import annotations

import math
import shutil
import time
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.optimize import differential_evolution
from scipy.spatial import cKDTree
from openpyxl import load_workbook, Workbook

try:
    from shapely.geometry import Polygon
except ImportError as exc:
    raise ImportError(
        "最终版需要 shapely。请执行：pip install -U shapely"
    ) from exc

# 0. 路径、常量、配置

BASE = Path(__file__).resolve().parent
DATA_FILE = BASE / "附件.xlsx"
RESULT2_TEMPLATE = BASE / "result2.xlsx"
RESULT3_TEMPLATE = BASE / "result3.xlsx"
OUT = BASE / "代码运行结果"
OUT.mkdir(exist_ok=True)

CFG = {
    # 地理参数
    "latitude_deg": 39.4,
    "longitude_deg": 98.5,
    "altitude_km": 3.0,

    # 题目参数
    "G0": 1.366,
    "tower_z": 80.0,
    "receiver_height": 8.0,
    "receiver_diameter": 7.0,
    "field_radius": 350.0,
    "tower_clearance": 100.0,
    "reflectivity": 0.92,

    # 问题1
    "q1_width": 6.0,
    "q1_height": 6.0,
    "q1_install_height": 4.0,

    # 问题2/3
    "target_power_MW": 60.0,
    "side_min": 2.0,
    "side_max": 8.0,
    "height_min": 2.0,
    "height_max": 6.0,
    "mirror_gap": 5.0,

    # 截断参数：题目/建模文档未完整提供
    # 这是为了让程序可以运行；若你们已有论文参数，请替换。
    "solar_half_angle_mrad": 4.65,
    "slope_error_mrad": 0.0,

    # Monte-Carlo
    "mc_rays_q1": 500,
    "mc_rays_final": 800,
    "mc_seed": 2026,
    "mc_batch": 512,

    # 阴影/遮挡
    "projection_grid": 5,
    "shadow_neighbor_radius": 25.0,

    # 优化
    "de_seed": 2026,
    "q2_maxiter": 2,
    "q2_popsize": 3,
    "q3_tower_maxiter": 0,
    "q3_tower_popsize": 0,

    # 先用快速模型筛选塔位，再用完整模型验证
    "fast_candidate_count": 1745,

    # 局部交换
    "local_swap_rounds": 1,

    # 收敛阈值
    "mc_tol": 0.01,
    "grid_tol": 0.01,

    # 运行速度控制：完整评价时只对代表性镜面做 Monte-Carlo，
    # 再用面积加权平均截断效率作用于整场；阴影遮挡采用KDTree快速几何近似。
    "truncation_sample_mirrors": 160,
    "convergence_sample_mirrors": 80,
    "convergence_states": [0, 15, 30, 45],
}


MONTHS = np.arange(1, 13)
TIMES = np.array([9.0, 10.5, 12.0, 13.5, 15.0])
N_STATES = 60

# 1. 数据结构

@dataclass
class Solution:
    problem: int
    tower_xy: np.ndarray
    positions: np.ndarray
    widths: np.ndarray
    heights: np.ndarray
    install_heights: np.ndarray
    eval_df: pd.DataFrame
    yearly: pd.DataFrame

    @property
    def area(self) -> float:
        return float(np.sum(self.widths * self.heights))

    @property
    def number(self) -> int:
        return len(self.positions)

    @property
    def power(self) -> float:
        return float(self.yearly.iloc[0]["年平均输出热功率(MW)"])

    @property
    def q(self) -> float:
        return self.power * 1000.0 / max(self.area, 1e-12)

# 2. 输入数据

def load_positions(path: Path) -> np.ndarray:
    df = pd.read_excel(path)

    if df.shape[1] < 2:
        raise ValueError("附件必须至少包含x、y两列。")

    xy = df.iloc[:, :2].apply(pd.to_numeric, errors="coerce").dropna().to_numpy(float)

    # 严格检查镜场区域
    r = np.linalg.norm(xy, axis=1)
    if np.any(r > CFG["field_radius"] + 1e-8):
        raise ValueError("发现候选镜面位于350m规划圆外。")

    return xy


def validate_positions(xy: np.ndarray) -> None:
    print(f"候选定日镜位置：{len(xy)}")
    print(f"x范围：[{xy[:,0].min():.3f}, {xy[:,0].max():.3f}]")
    print(f"y范围：[{xy[:,1].min():.3f}, {xy[:,1].max():.3f}]")
    r = np.linalg.norm(xy, axis=1)
    print(f"距场心最小/最大：{r.min():.3f} / {r.max():.3f} m")


# 3. 太阳模型

def day_from_spring(month: int) -> int:

    month_starts = np.array(
        [0, 31, 61, 92, 122, 153, 183, 214, 245, 275, 306, 336]
    )
    doy = month_starts[month - 1] + 21
    return int(doy - 80)


def solar_state(month: int, st: float) -> dict:
    phi = math.radians(CFG["latitude_deg"])
    D = day_from_spring(month)

    delta = math.asin(
        math.sin(2 * math.pi * D / 365.0)
        * math.sin(math.radians(23.45))
    )

    omega = math.pi / 12.0 * (st - 12.0)

    sin_alpha = (
        math.cos(delta) * math.cos(phi) * math.cos(omega)
        + math.sin(delta) * math.sin(phi)
    )
    sin_alpha = np.clip(sin_alpha, -1.0, 1.0)
    alpha = math.asin(sin_alpha)

    # 东-北-天顶坐标
    east = -math.cos(delta) * math.sin(omega)
    north = (
        math.sin(delta) * math.cos(phi)
        - math.cos(delta) * math.sin(phi) * math.cos(omega)
    )
    up = sin_alpha

    s = np.array([east, north, up], dtype=float)
    s /= np.linalg.norm(s)

    gamma = math.atan2(east, north)

    return {
        "month": month,
        "time": st,
        "D": D,
        "alpha_deg": math.degrees(alpha),
        "gamma_deg": math.degrees(gamma),
        "s": s,
    }


def dni_from_alpha(alpha_deg: float) -> float:
    if alpha_deg <= 0:
        return 0.0

    H = CFG["altitude_km"]
    G0 = CFG["G0"]

    a = 0.4237 - 0.00821 * (6.0 - H) ** 2
    b = 0.5055 + 0.00595 * (6.5 - H) ** 2
    c = 0.2711 + 0.01858 * (2.5 - H) ** 2

    sa = math.sin(math.radians(alpha_deg))
    sa = max(sa, 1e-10)

    return G0 * (a + b * math.exp(-c / sa))


def build_states() -> pd.DataFrame:
    rows = []

    for m in MONTHS:
        for st in TIMES:
            s = solar_state(int(m), float(st))
            rows.append({
                "month": s["month"],
                "time": s["time"],
                "D": s["D"],
                "alpha_deg": s["alpha_deg"],
                "gamma_deg": s["gamma_deg"],
                "sx": s["s"][0],
                "sy": s["s"][1],
                "sz": s["s"][2],
                "DNI_kW_m2": dni_from_alpha(s["alpha_deg"]),
            })

    return pd.DataFrame(rows)


STATES = build_states()


# 4. 镜面几何

def mirror_geometry(
    positions: np.ndarray,
    install_heights: np.ndarray,
    tower_xy: np.ndarray,
    sun: np.ndarray,
) -> dict:
    M = np.column_stack([
        positions[:, 0],
        positions[:, 1],
        install_heights
    ])

    T = np.array([
        tower_xy[0],
        tower_xy[1],
        CFG["tower_z"]
    ])

    # 镜面 -> 集热器中心
    R = T[None, :] - M
    d = np.linalg.norm(R, axis=1)
    r = R / d[:, None]

    # 镜面 -> 太阳
    s = np.broadcast_to(sun, M.shape).copy()

    # 法向：角平分线
    h = s + r
    hn = np.linalg.norm(h, axis=1)

    normal = np.zeros_like(h)
    good = hn > 1e-12
    normal[good] = h[good] / hn[good, None]

    # 余弦效率
    eta_cos = np.clip(np.sum(s * normal, axis=1), 0.0, 1.0)

    # 大气透射
    eta_at = (
        0.99321
        - 0.0001176 * d
        + 1.97e-8 * d**2
    )
    eta_at = np.clip(eta_at, 0.0, 1.0)

    return {
        "M": M,
        "r": r,
        "normal": normal,
        "eta_cos": eta_cos,
        "eta_at": eta_at,
        "distance": d,
    }

# 5. 镜面局部坐标与投影

def mirror_basis(normal: np.ndarray) -> tuple[np.ndarray, np.ndarray]:

    z = np.array([0.0, 0.0, 1.0])

    u = np.array([-normal[1], normal[0], 0.0])
    nu = np.linalg.norm(u)

    if nu < 1e-12:
        u = np.array([1.0, 0.0, 0.0])
    else:
        u /= nu

    v = np.cross(normal, u)
    v /= np.linalg.norm(v)

    # 保证v方向具有正z分量
    if v[2] < 0:
        v = -v

    return u, v


def mirror_corners(
    center: np.ndarray,
    normal: np.ndarray,
    width: float,
    height: float,
) -> np.ndarray:
    u, v = mirror_basis(normal)

    return np.array([
        center - width / 2 * u - height / 2 * v,
        center + width / 2 * u - height / 2 * v,
        center + width / 2 * u + height / 2 * v,
        center - width / 2 * u + height / 2 * v,
    ])


def projection_basis(direction: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    d = direction / np.linalg.norm(direction)

    ref = np.array([0.0, 0.0, 1.0])
    if abs(np.dot(d, ref)) > 0.95:
        ref = np.array([1.0, 0.0, 0.0])

    e1 = np.cross(d, ref)
    e1 /= np.linalg.norm(e1)

    e2 = np.cross(d, e1)
    e2 /= np.linalg.norm(e2)

    return e1, e2


def project_polygon(points3d: np.ndarray, direction: np.ndarray) -> Polygon:
    e1, e2 = projection_basis(direction)
    xy = np.column_stack([
        points3d @ e1,
        points3d @ e2
    ])

    poly = Polygon(xy)
    if not poly.is_valid:
        poly = poly.buffer(0)

    return poly

# 6. 阴影与遮挡

def shadow_block_efficiency(
    positions: np.ndarray,
    install_heights: np.ndarray,
    widths: np.ndarray,
    heights: np.ndarray,
    tower_xy: np.ndarray,
    sun: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:

    n = len(positions)
    geom = mirror_geometry(positions, install_heights, tower_xy, sun)
    M = geom["M"]
    R = geom["r"]
    tree = cKDTree(positions)

    eta_shadow = np.ones(n)
    eta_block = np.ones(n)
    diag = 0.5 * np.sqrt(widths * widths + heights * heights)
    radius = CFG["shadow_neighbor_radius"]

    for i in range(n):
        neighbors = tree.query_ball_point(positions[i], radius)
        if len(neighbors) <= 1:
            continue
        js = np.asarray([j for j in neighbors if j != i], dtype=int)
        delta = M[js] - M[i]
        dist3 = np.linalg.norm(delta, axis=1)

        # 阴影：邻镜必须处于太阳入射方向的上游。
        sun_proj = delta @ sun
        upstream = sun_proj > 0.0
        if np.any(upstream):
            d = delta[upstream]
            jj = js[upstream]
            proj = np.abs(d @ sun)
            lateral = np.sqrt(np.maximum(dist3[upstream] ** 2 - proj ** 2, 0.0))
            threshold = diag[i] + diag[jj] + CFG["mirror_gap"] * 0.25
            overlap = np.clip(1.0 - lateral / np.maximum(threshold, 1e-9), 0.0, 1.0)
            if overlap.size:
                eta_shadow[i] = 1.0 - float(np.max(overlap))

        # 遮挡：邻镜位于镜面到吸收塔的方向上。
        ray = R[i]
        block_proj = delta @ ray
        forward = block_proj > 0.0
        if np.any(forward):
            d = delta[forward]
            jj = js[forward]
            rr = np.linalg.norm(ray)
            proj = np.abs(d @ (ray / rr))
            lateral = np.sqrt(np.maximum(dist3[forward] ** 2 - proj ** 2, 0.0))
            threshold = diag[i] + diag[jj] + CFG["mirror_gap"] * 0.25
            overlap = np.clip(1.0 - lateral / np.maximum(threshold, 1e-9), 0.0, 1.0)
            if overlap.size:
                eta_block[i] = 1.0 - float(np.max(overlap))

    return np.clip(eta_shadow, 0.0, 1.0), np.clip(eta_block, 0.0, 1.0)

# 7. 截断效率 Monte-Carlo

def cylinder_hit(
    origins: np.ndarray,
    directions: np.ndarray,
    tower_xy: np.ndarray,
) -> np.ndarray:

    cx, cy = tower_xy
    R = CFG["receiver_diameter"] / 2.0
    z0 = CFG["tower_z"] - CFG["receiver_height"] / 2.0
    z1 = CFG["tower_z"] + CFG["receiver_height"] / 2.0

    ox = origins[:, 0] - cx
    oy = origins[:, 1] - cy
    dx = directions[:, 0]
    dy = directions[:, 1]

    A = dx * dx + dy * dy
    B = 2.0 * (ox * dx + oy * dy)
    C = ox * ox + oy * oy - R * R

    disc = B * B - 4.0 * A * C
    valid = disc >= 0

    t = np.full(len(origins), np.inf)

    safe = valid & (A > 1e-14)

    if np.any(safe):
        root = np.sqrt(np.maximum(disc[safe], 0.0))
        t1 = (-B[safe] - root) / (2.0 * A[safe])
        t2 = (-B[safe] + root) / (2.0 * A[safe])

        positive1 = t1 > 0
        t[safe] = np.where(positive1, t1, t2)

    valid &= t > 0

    z_hit = origins[:, 2] + t * directions[:, 2]

    valid &= z_hit >= z0
    valid &= z_hit <= z1

    return valid


def truncation_efficiency(
    positions: np.ndarray,
    install_heights: np.ndarray,
    tower_xy: np.ndarray,
    sun: np.ndarray,
    rays: int,
    seed: int,
    sample_indices: np.ndarray | None = None,
) -> np.ndarray:

    geom = mirror_geometry(positions, install_heights, tower_xy, sun)
    M = geom["M"]
    R = geom["r"]
    n = len(positions)

    if sample_indices is None:
        max_n = int(CFG.get("truncation_sample_mirrors", 160))
        if n <= max_n:
            sample_indices = np.arange(n, dtype=int)
        else:
            sample_indices = np.linspace(0, n - 1, max_n, dtype=int)
    else:
        sample_indices = np.asarray(sample_indices, dtype=int)

    sigma = math.sqrt(CFG["solar_half_angle_mrad"] ** 2 + CFG["slope_error_mrad"] ** 2) * 1e-3
    rng = np.random.default_rng(seed)
    sample_values = np.empty(len(sample_indices), dtype=float)

    for kk, i in enumerate(sample_indices):
        ri = R[i]
        ref = np.array([0.0, 0.0, 1.0])
        if abs(np.dot(ref, ri)) > 0.95:
            ref = np.array([1.0, 0.0, 0.0])
        e1 = np.cross(ri, ref)
        e1 /= np.linalg.norm(e1)
        e2 = np.cross(ri, e1)
        e2 /= np.linalg.norm(e2)

        a = rng.normal(0.0, sigma, rays)
        b = rng.normal(0.0, sigma, rays)
        directions = ri[None, :] + a[:, None] * e1[None, :] + b[:, None] * e2[None, :]
        directions /= np.linalg.norm(directions, axis=1)[:, None]
        origins = np.broadcast_to(M[i], (rays, 3))
        sample_values[kk] = np.mean(cylinder_hit(origins, directions, tower_xy))

    value = float(np.mean(sample_values)) if len(sample_values) else 1.0
    return np.full(n, np.clip(value, 0.0, 1.0), dtype=float)

# 8. 单时刻 / 60时刻镜场评价

def evaluate_one_state(
    positions: np.ndarray,
    widths: np.ndarray,
    heights: np.ndarray,
    install_heights: np.ndarray,
    tower_xy: np.ndarray,
    state: pd.Series,
    full: bool = True,
    rays: int = 3000,
    seed: int = 2026,
) -> dict:
    sun = np.array([
        state["sx"],
        state["sy"],
        state["sz"]
    ])

    DNI = float(state["DNI_kW_m2"])

    geom = mirror_geometry(
        positions,
        install_heights,
        tower_xy,
        sun
    )

    eta_cos = geom["eta_cos"]
    eta_at = geom["eta_at"]

    if full:
        eta_shadow, eta_block = shadow_block_efficiency(
            positions,
            install_heights,
            widths,
            heights,
            tower_xy,
            sun
        )
        eta_sb = eta_shadow * eta_block

        eta_trunc = truncation_efficiency(
            positions,
            install_heights,
            tower_xy,
            sun,
            rays,
            seed
        )
    else:
        eta_shadow = np.ones(len(positions))
        eta_block = np.ones(len(positions))
        eta_sb = np.ones(len(positions))
        eta_trunc = np.ones(len(positions))

    eta = (
        eta_sb
        * eta_cos
        * eta_at
        * eta_trunc
        * CFG["reflectivity"]
    )

    area = widths * heights
    total_area = np.sum(area)

    power_kW = DNI * np.sum(area * eta)

    # 面积加权效率
    avg_optical = np.average(
        eta,
        weights=area
    )
    avg_cos = np.average(
        eta_cos,
        weights=area
    )
    avg_shadow = np.average(
        eta_sb,
        weights=area
    )
    avg_trunc = np.average(
        eta_trunc,
        weights=area
    )

    return {
        "month": int(state["month"]),
        "time": float(state["time"]),
        "DNI_kW_m2": DNI,
        "eta_optical_mean": avg_optical,
        "eta_cos_mean": avg_cos,
        "eta_shadow_mean": avg_shadow,
        "eta_trunc_mean": avg_trunc,
        "power_MW": power_kW / 1000.0,
        "unit_power_kW_m2": power_kW / max(total_area, 1e-12),
    }


def evaluate_field(
    positions: np.ndarray,
    widths: np.ndarray,
    heights: np.ndarray,
    install_heights: np.ndarray,
    tower_xy: np.ndarray,
    full: bool,
    rays: int,
    seed: int,
) -> pd.DataFrame:
    rows = []

    for k, (_, state) in enumerate(STATES.iterrows()):
        rows.append(
            evaluate_one_state(
                positions,
                widths,
                heights,
                install_heights,
                tower_xy,
                state,
                full=full,
                rays=rays,
                seed=seed + k
            )
        )

    return pd.DataFrame(rows)


def summarize_year(eval_df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame([{
        "年平均光学效率": eval_df["eta_optical_mean"].mean(),
        "年平均余弦效率": eval_df["eta_cos_mean"].mean(),
        "年平均阴影遮挡效率": eval_df["eta_shadow_mean"].mean(),
        "年平均截断效率": eval_df["eta_trunc_mean"].mean(),
        "年平均输出热功率(MW)": eval_df["power_MW"].mean(),
        "单位镜面面积年平均输出热功率(kW/m2)":
            eval_df["unit_power_kW_m2"].mean(),
    }])


def summarize_month(eval_df: pd.DataFrame) -> pd.DataFrame:
    out = (
        eval_df
        .groupby("month")
        .agg({
            "eta_optical_mean": "mean",
            "eta_cos_mean": "mean",
            "eta_shadow_mean": "mean",
            "eta_trunc_mean": "mean",
            "power_MW": "mean",
            "unit_power_kW_m2": "mean",
        })
        .reset_index()
    )

    out.columns = [
        "月份",
        "平均光学效率",
        "平均余弦效率",
        "平均阴影遮挡效率",
        "平均截断效率",
        "平均输出热功率(MW)",
        "单位面积平均输出热功率(kW/m2)",
    ]

    return out

# 9. 问题1

def solve_q1(positions: np.ndarray) -> Solution:
    print("\n问题1")

    n = len(positions)

    widths = np.full(n, 6.0)
    heights = np.full(n, 6.0)
    install = np.full(n, 4.0)
    tower = np.array([0.0, 0.0])

    t0 = time.time()

    eval_df = evaluate_field(
        positions,
        widths,
        heights,
        install,
        tower,
        full=True,
        rays=CFG["mc_rays_q1"],
        seed=CFG["mc_seed"]
    )

    yearly = summarize_year(eval_df)

    print(f"功率 {yearly.iloc[0]['年平均输出热功率(MW)']:.3f} MW")

    return Solution(
        problem=1,
        tower_xy=tower,
        positions=positions,
        widths=widths,
        heights=heights,
        install_heights=install,
        eval_df=eval_df,
        yearly=yearly
    )

# 10. 问题2：候选镜面评分

def mirror_score_q2(
    positions: np.ndarray,
    width: float,
    height: float,
    install_height: float,
    tower_xy: np.ndarray,
    state_indices: np.ndarray | None = None,
) -> np.ndarray:

    area = width * height
    score = np.zeros(len(positions))
    if state_indices is None:
        state_indices = np.arange(len(STATES), dtype=int)

    for si in state_indices:
        state = STATES.iloc[int(si)]
        sun = np.array([
            state["sx"],
            state["sy"],
            state["sz"]
        ])

        geom = mirror_geometry(
            positions,
            np.full(len(positions), install_height),
            tower_xy,
            sun
        )

        score += (
            state["DNI_kW_m2"]
            * geom["eta_cos"]
            * geom["eta_at"]
            * CFG["reflectivity"]
        )

    return area * score / max(len(state_indices), 1)


def feasible_greedy_q2(
    positions: np.ndarray,
    width: float,
    height: float,
    install_height: float,
    tower_xy: np.ndarray,
    state_indices: np.ndarray | None = None,
) -> np.ndarray:

    score = mirror_score_q2(
        positions,
        width,
        height,
        install_height,
        tower_xy,
        state_indices=state_indices
    )

    order = np.argsort(-score)

    tree = cKDTree(positions)

    selected = []
    selected_set = set()
    approx_power_kW = 0.0

    min_dist = width + CFG["mirror_gap"]

    for idx in order:
        p = positions[idx]

        if np.linalg.norm(
            p - tower_xy
        ) < CFG["tower_clearance"]:
            continue

        if selected:
            neighbors = tree.query_ball_point(
                p,
                min_dist
            )
            if any(j in selected_set for j in neighbors):
                continue

        selected.append(int(idx))
        selected_set.add(int(idx))

        # 快速功率上界，采用累计值，避免每次重复求和整个已选集合
        approx_power_kW += float(score[idx])

        if approx_power_kW >= CFG["target_power_MW"] * 1000:
            break

    return np.array(selected, dtype=int)


def q2_fast_objective(
    x: np.ndarray,
    positions: np.ndarray,
) -> tuple[float, dict]:
    tx, ty, W, L, H = x

    tower = np.array([tx, ty])

    # 塔位约束
    if np.linalg.norm(tower) > 100:
        return 1e9, {}

    # W >= L
    if W < L:
        return 1e8 + (L - W) * 1e5, {}

    selected = feasible_greedy_q2(
        positions,
        W,
        L,
        H,
        tower,
        state_indices=np.arange(0, N_STATES, 5, dtype=int)
    )

    if len(selected) == 0:
        return 1e9, {}

    p = positions[selected]

    widths = np.full(len(p), W)
    heights = np.full(len(p), L)
    install = np.full(len(p), H)

    # 快速模型只算余弦、大气、反射率；优化阶段使用12个代表时刻
    fast_idx = np.arange(0, N_STATES, 5, dtype=int)
    power_values = []
    for si in fast_idx:
        state = STATES.iloc[int(si)]
        sun = np.array([state["sx"], state["sy"], state["sz"]])
        geom = mirror_geometry(p, install, tower, sun)
        eta = geom["eta_cos"] * geom["eta_at"] * CFG["reflectivity"]
        power_values.append(float(state["DNI_kW_m2"] * np.sum((widths * heights) * eta) / 1000.0))
    power = float(np.mean(power_values))
    area = len(p) * W * L

    q = power * 1000 / max(area, 1e-12)

    if power < CFG["target_power_MW"]:
        penalty = (
            CFG["target_power_MW"] - power
        ) ** 2 * 20

        return -q + penalty, {
            "tower": tower,
            "W": W,
            "L": L,
            "H": H,
            "selected": selected,
            "fast_power": power,
            "q": q
        }

    return -q, {
        "tower": tower,
        "W": W,
        "L": L,
        "H": H,
        "selected": selected,
        "fast_power": power,
        "q": q
    }


def solve_q2(positions: np.ndarray) -> Solution:
    print("\n问题2", flush=True)
    t0 = time.time()

    # 固定随机种子的轻量全局搜索，替代高开销差分进化。
    # 先用少量确定性点覆盖边界，再用固定随机点补充搜索。
    rng = np.random.default_rng(CFG["de_seed"])
    candidates = [
        [0., 0., 6., 6., 4.],
        [0., 0., 8., 6., 4.],
        [0., 0., 8., 8., 6.],
        [40., 0., 6., 6., 4.],
        [-40., 0., 6., 6., 4.],
        [0., 40., 6., 6., 4.],
        [0., -40., 6., 6., 4.],
        [60., 0., 8., 6., 5.],
        [0., 60., 8., 6., 5.],
    ]
    for _ in range(30):
        angle = rng.uniform(0.0, 2.0 * math.pi)
        radius = math.sqrt(rng.uniform(0.0, 100.0 ** 2))
        tx, ty = radius * math.cos(angle), radius * math.sin(angle)
        L = rng.uniform(2.0, 8.0)
        W = rng.uniform(L, 8.0)
        H = rng.uniform(2.0, 6.0)
        candidates.append([tx, ty, W, L, H])

    best = {"value": np.inf, "info": None}
    for k, x in enumerate(candidates, 1):
        value, info = q2_fast_objective(np.asarray(x, dtype=float), positions)
        if info and value < best["value"]:
            best["value"] = value
            best["info"] = info
        if k % 10 == 0 or k == len(candidates):
            print(f"问题2搜索 {k}/{len(candidates)}", flush=True)

    if best["info"] is None:
        raise RuntimeError("问题2没有找到可行初始方案。")

    info = best["info"]
    selected = info["selected"]
    p = positions[selected]
    widths = np.full(len(p), info["W"])
    heights = np.full(len(p), info["L"])
    install = np.full(len(p), info["H"])

    print(f"问题2最终精算，镜面数 {len(p)}", flush=True)
    eval_df = evaluate_field(
        p, widths, heights, install, info["tower"],
        full=True, rays=CFG["mc_rays_final"], seed=CFG["mc_seed"]
    )
    yearly = summarize_year(eval_df)
    print(f"功率 {yearly.iloc[0]['年平均输出热功率(MW)']:.3f} MW", flush=True)

    return Solution(
        problem=2,
        tower_xy=info["tower"],
        positions=p,
        widths=widths,
        heights=heights,
        install_heights=install,
        eval_df=eval_df,
        yearly=yearly
    )

# 11. 问题3

def q3_initial_dimensions(
    positions: np.ndarray,
    tower_xy: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:

    d = np.linalg.norm(
        positions - tower_xy[None, :],
        axis=1
    )

    # 距离越远，允许更大的镜面；
    # 维持2~8m且W>=L。
    W = np.clip(
        2.0 + 0.014 * np.maximum(d - 100, 0),
        2.0,
        8.0
    )

    L = np.clip(
        0.85 * W,
        2.0,
        8.0
    )

    H = np.full(
        len(positions),
        4.0
    )

    return W, L, H


def q3_greedy(positions, tower_xy, fast_target_MW=None):

    if fast_target_MW is None:
        fast_target_MW = CFG.get("q3_fast_target_MW", 70.0)

    unit_score = mirror_score_q2(positions, 1.0, 1.0, 4.0, tower_xy, state_indices=np.arange(0, N_STATES, 5, dtype=int))
    W0, L0, H0 = q3_initial_dimensions(positions, tower_xy)
    score = unit_score * W0 * L0
    order = np.argsort(-score)
    tree = cKDTree(positions)

    selected, widths, heights, install = [], [], [], []
    selected_set = set()
    selected_pos = {}
    approx_power_kW = 0.0
    target_kW = fast_target_MW * 1000.0

    for idx in order:
        p = positions[idx]
        if np.linalg.norm(p - tower_xy) < CFG["tower_clearance"]:
            continue

        wi, li, hi = float(W0[idx]), float(L0[idx]), float(H0[idx])
        conflict = False
        neighbors = tree.query_ball_point(p, CFG["side_max"] + CFG["mirror_gap"])
        for j in neighbors:
            if j not in selected_set:
                continue
            k = selected_pos[j]
            required = max(wi, widths[k]) + CFG["mirror_gap"]
            if np.linalg.norm(p - positions[j]) < required - 1e-10:
                conflict = True
                break
        if conflict:
            continue

        selected.append(int(idx))
        selected_set.add(int(idx))
        selected_pos[int(idx)] = len(selected) - 1
        widths.append(wi); heights.append(li); install.append(hi)
        approx_power_kW += unit_score[idx] * wi * li
        if approx_power_kW >= target_kW:
            break

    return (np.asarray(selected, dtype=int), np.asarray(widths),
            np.asarray(heights), np.asarray(install))


def improve_q3_dimensions(sol_positions, tower_xy, widths, heights, install):

    W, L, H = widths.copy(), heights.copy(), install.copy()
    if len(W) == 0:
        return W, L, H

    unit_score = mirror_score_q2(sol_positions, 1.0, 1.0, 4.0, tower_xy, state_indices=np.arange(0, N_STATES, 5, dtype=int))
    target_kW = CFG.get("q3_fast_target_MW", 70.0) * 1000.0
    levels = np.array([2., 3., 4., 5., 6., 7., 8.])
    hlevels = np.array([2., 3., 4., 5., 6.])

    for _ in range(2):
        total_power = float(np.sum(unit_score * W * L))
        order = np.argsort(unit_score)
        for i in order:
            ow, ol, oh = W[i], L[i], H[i]
            old_c = unit_score[i] * ow * ol
            best = (ow, ol, oh, total_power)
            wc = levels[np.abs(levels-ow) <= 1.000001]
            for nw in wc:
                lc = levels[(levels >= 2) & (levels <= nw + 1e-12) & (np.abs(levels-ol) <= 1.000001)]
                hc = hlevels[np.abs(hlevels-oh) <= 1.000001]
                for nl in lc:
                    for nh in hc:
                        npower = total_power - old_c + unit_score[i] * nw * nl
                        if npower >= target_kW:
                            area_new = nw * nl
                            area_best = best[0] * best[1]
                            if area_new < area_best - 1e-12 or (abs(area_new-area_best)<1e-12 and npower>best[3]):
                                best = (nw, nl, nh, npower)
            W[i], L[i], H[i], total_power = best
    return W, L, H


def _q3_fast_candidate(positions, tower):
    selected, W, L, H = q3_greedy(positions, tower)
    if len(selected) == 0:
        return None
    p = positions[selected]
    W, L, H = improve_q3_dimensions(p, tower, W, L, H)
    unit_score = mirror_score_q2(p, 1.0, 1.0, 4.0, tower, state_indices=np.arange(0, N_STATES, 5, dtype=int))
    power = float(np.sum(unit_score * W * L) / 1000.0)
    area = float(np.sum(W * L))
    return {"tower": np.asarray(tower), "selected": selected, "W": W, "L": L, "H": H,
            "fast_power": power, "q": power*1000/max(area,1e-12)}


def solve_q3(positions):
    print("\n问题3开始")
    t0 = time.time()

    tower_candidates = np.array([[0.,0.],[40.,0.],[-40.,0.],[0.,40.],[0.,-40.]])
    best = None
    for k, tower in enumerate(tower_candidates, 1):
        print(f"问题3粗筛 {k}/{len(tower_candidates)}", flush=True)
        c = _q3_fast_candidate(positions, tower)
        if c is None:
            print("    无可行镜面", flush=True); continue
        print(f"  镜面 {len(c['selected'])}，快速功率 {c['fast_power']:.3f} MW，单位面积 {c['q']:.3f} kW/m²", flush=True)
        if best is None or c['q'] > best['q']:
            best = c

    if best is None:
        raise RuntimeError("问题3没有找到可行初始方案。")

    center = best['tower'].copy()
    offsets = [-10., 0., 10.]
    local = []
    for dx in offsets:
        for dy in offsets:
            t = center + np.array([dx,dy])
            if np.linalg.norm(t) <= 100.:
                local.append(t)

    print(f"问题3局部搜索，共 {len(local)} 个塔位", flush=True)
    for k,tower in enumerate(local,1):
        c = _q3_fast_candidate(positions, tower)
        if c is not None and ((c['fast_power'] >= best['fast_power'] and c['q'] >= best['q']) or (best['fast_power'] < 70 and c['fast_power'] > best['fast_power'])):
            best = c
        if k % 10 == 0 or k == len(local):
            print(f"  已完成 {k}/{len(local)}，当前最好功率 {best['fast_power']:.3f} MW", flush=True)

    tower, selected = best['tower'], best['selected']
    p = positions[selected]
    W, L, H = best['W'], best['L'], best['H']

    print(f"问题3最终精算，镜面数 {len(p)}", flush=True)
    print("开始最终60时刻光学计算", flush=True)
    eval_df = evaluate_field(p, W, L, H, tower, full=True, rays=CFG['mc_rays_final'], seed=CFG['mc_seed'])
    yearly = summarize_year(eval_df)
    print("\n问题3最终方案：")
    print(f"塔坐标：{tower}")
    print(f"镜面数：{len(p)}")
    print(f"总面积：{np.sum(W*L):.6f} m²")
    print(f"功率 {yearly.iloc[0]['年平均输出热功率(MW)']:.3f} MW")

    return Solution(problem=3, tower_xy=tower, positions=p, widths=W, heights=L, install_heights=H, eval_df=eval_df, yearly=yearly)

# 12. Monte-Carlo收敛检验

def mc_convergence(solution: Solution) -> pd.DataFrame:
    K_list = [200, 400, 800]
    idx = np.linspace(0, solution.number - 1,
                      min(CFG["convergence_sample_mirrors"], solution.number), dtype=int)
    state_indices = CFG["convergence_states"]
    rows = []

    sample_pos = solution.positions[idx]
    sample_h = solution.install_heights[idx]
    for K in K_list:
        vals1 = []
        vals2 = []
        for si in state_indices:
            state = STATES.iloc[int(si)]
            sun = np.array([state["sx"], state["sy"], state["sz"]])
            e1 = truncation_efficiency(sample_pos, sample_h, solution.tower_xy, sun, K, CFG["mc_seed"] + int(si), np.arange(len(idx)))
            e2 = truncation_efficiency(sample_pos, sample_h, solution.tower_xy, sun, 2*K, CFG["mc_seed"] + int(si), np.arange(len(idx)))
            vals1.append(float(np.mean(e1)))
            vals2.append(float(np.mean(e2)))
        a = float(np.mean(vals1))
        b = float(np.mean(vals2))
        delta = abs(a-b) / max(abs(b), 1e-12)
        rows.append({"K": K, "2K": 2*K, "eta_trunc_K": a, "eta_trunc_2K": b,
                     "相对变化率": delta, "是否达到1%阈值": delta < CFG["mc_tol"]})
    return pd.DataFrame(rows)

# 13. 网格收敛检验

def grid_convergence(solution: Solution) -> pd.DataFrame:
    return pd.DataFrame([{
        "网格": "快速KDTree几何法",
        "平均阴影遮挡效率": float(solution.eval_df["eta_shadow_mean"].mean()),
        "相对变化率": np.nan,
        "说明": "当前快速几何遮挡算法无投影网格离散参数，跳过重复网格计算"
    }])

# 14. Excel结果输出

def write_result_excel(
    template: Path,
    output: Path,
    solution: Solution,
) -> None:

    if template.exists():
        shutil.copy2(template, output)
        wb = load_workbook(output)
        ws = wb.worksheets[0]

        # 保留第1行表头
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        wb = Workbook()
        ws = wb.active

        headers = [
            "吸收塔x坐标 (m)",
            "吸收塔y坐标 (m)",
            "定日镜序号",
            "定日镜宽度 (m)",
            "定日镜高度 (m)",
            "定日镜x坐标 (m)",
            "定日镜y坐标 (m)",
            "定日镜z坐标 (m)"
        ]

        ws.append(headers)

    for i in range(solution.number):
        ws.append([
            float(solution.tower_xy[0]),
            float(solution.tower_xy[1]),
            i + 1,
            float(solution.widths[i]),
            float(solution.heights[i]),
            float(solution.positions[i, 0]),
            float(solution.positions[i, 1]),
            float(solution.install_heights[i])
        ])

    # 增加结果汇总页，不破坏原模板Sheet1
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    sm = wb.create_sheet("Summary")

    for c, name in enumerate(solution.yearly.columns, 1):
        sm.cell(1, c, name)

    for c, value in enumerate(
        solution.yearly.iloc[0].tolist(),
        1
    ):
        sm.cell(
            2,
            c,
            float(value)
            if isinstance(value, (float, np.floating))
            else value
        )

    sm["A4"] = "吸收塔x坐标(m)"
    sm["B4"] = float(solution.tower_xy[0])

    sm["A5"] = "吸收塔y坐标(m)"
    sm["B5"] = float(solution.tower_xy[1])

    sm["A6"] = "定日镜总数"
    sm["B6"] = solution.number

    sm["A7"] = "定日镜总面积(m2)"
    sm["B7"] = solution.area

    sm["A8"] = "年平均输出热功率(MW)"
    sm["B8"] = solution.power

    sm["A9"] = "单位面积年平均输出热功率(kW/m2)"
    sm["B9"] = solution.q

    wb.save(output)


def save_all_tables(
    name: str,
    solution: Solution,
) -> None:
    month = summarize_month(solution.eval_df)

    solution.eval_df.to_csv(
        OUT / f"{name}_60时刻明细.csv",
        index=False,
        encoding="utf-8-sig"
    )

    month.to_csv(
        OUT / f"{name}_Table1.csv",
        index=False,
        encoding="utf-8-sig"
    )

    solution.yearly.to_csv(
        OUT / f"{name}_Table2.csv",
        index=False,
        encoding="utf-8-sig"
    )

    mirrors = pd.DataFrame({
        "定日镜序号": np.arange(
            1,
            solution.number + 1
        ),
        "吸收塔x坐标(m)": solution.tower_xy[0],
        "吸收塔y坐标(m)": solution.tower_xy[1],
        "定日镜宽度(m)": solution.widths,
        "定日镜高度(m)": solution.heights,
        "定日镜x坐标(m)": solution.positions[:, 0],
        "定日镜y坐标(m)": solution.positions[:, 1],
        "定日镜z坐标(m)": solution.install_heights,
        "镜面面积(m2)": solution.widths * solution.heights
    })

    mirrors.to_csv(
        OUT / f"{name}_镜面明细.csv",
        index=False,
        encoding="utf-8-sig"
    )

# 15. 论文图

def plot_q1(solution: Solution):
    month = summarize_month(solution.eval_df)

    plt.figure(figsize=(10, 5))

    plt.plot(
        month["月份"],
        month["平均光学效率"],
        marker="o",
        label="Optical efficiency"
    )

    plt.plot(
        month["月份"],
        month["平均余弦效率"],
        marker="s",
        label="Cosine efficiency"
    )

    plt.plot(
        month["月份"],
        month["平均阴影遮挡效率"],
        marker="^",
        label="Shadow/blocking efficiency"
    )

    plt.plot(
        month["月份"],
        month["平均截断效率"],
        marker="d",
        label="Truncation efficiency"
    )

    plt.xlabel("Month")
    plt.ylabel("Efficiency")
    plt.xticks(np.arange(1, 13))
    plt.grid(alpha=0.25)
    plt.legend()
    plt.tight_layout()

    plt.savefig(
        OUT / "Q1_效率曲线.png",
        dpi=300
    )

    plt.close()

    plt.figure(figsize=(10, 5))

    plt.plot(
        month["月份"],
        month["单位面积平均输出热功率(kW/m2)"],
        marker="o"
    )

    plt.xlabel("Month")
    plt.ylabel("Average output power per mirror area (kW/m2)")
    plt.xticks(np.arange(1, 13))
    plt.grid(alpha=0.25)
    plt.tight_layout()

    plt.savefig(
        OUT / "Q1_单位面积功率.png",
        dpi=300
    )

    plt.close()


def plot_layout(
    solution: Solution,
    filename: str
):
    fig, ax = plt.subplots(figsize=(8, 8))

    # 全部规划区
    field = plt.Circle(
        (0, 0),
        CFG["field_radius"],
        fill=False,
        linewidth=1.5
    )

    clearance = plt.Circle(
        tuple(solution.tower_xy),
        CFG["tower_clearance"],
        fill=False,
        linestyle="--"
    )

    ax.add_patch(field)
    ax.add_patch(clearance)

    ax.scatter(
        solution.positions[:, 0],
        solution.positions[:, 1],
        s=10,
        alpha=0.65
    )

    ax.scatter(
        [solution.tower_xy[0]],
        [solution.tower_xy[1]],
        marker="*",
        s=200,
        label="Tower"
    )

    ax.set_aspect("equal")
    ax.set_xlabel("x / m")
    ax.set_ylabel("y / m")
    ax.grid(alpha=0.2)
    ax.legend()

    plt.tight_layout()

    plt.savefig(
        OUT / filename,
        dpi=300
    )

    plt.close()


def plot_q2_q3(q2: Solution, q3: Solution):
    names = ["Q2", "Q3"]
    powers = [q2.power, q3.power]
    qs = [q2.q, q3.q]

    x = np.arange(2)

    plt.figure(figsize=(8, 5))
    plt.bar(x, powers)
    plt.xticks(x, names)
    plt.ylabel("Annual average thermal power (MW)")
    plt.grid(axis="y", alpha=0.25)
    plt.tight_layout()
    plt.savefig(
        OUT / "Q2_Q3_年平均功率对比.png",
        dpi=300
    )
    plt.close()

    plt.figure(figsize=(8, 5))
    plt.bar(x, qs)
    plt.xticks(x, names)
    plt.ylabel("Annual average output per mirror area (kW/m2)")
    plt.grid(axis="y", alpha=0.25)
    plt.tight_layout()
    plt.savefig(
        OUT / "Q2_Q3_单位面积指标对比.png",
        dpi=300
    )
    plt.close()

# 16. 主程序

def main():
    total_start = time.time()

    print("A题 定日镜场优化设计")
    print("2023 高教社杯 A题 —— 定日镜场优化设计")
    print("FINAL COMPETITION VERSION")
    print("A题 定日镜场优化设计")

    if not DATA_FILE.exists():
        raise FileNotFoundError(
            f"缺少：{DATA_FILE.name}"
        )

    positions = load_positions(
        DATA_FILE
    )

    validate_positions(
        positions
    )

    STATES.to_csv(
        OUT / "太阳位置_DNI_60时刻.csv",
        index=False,
        encoding="utf-8-sig"
    )

    q1 = solve_q1(
        positions
    )

    save_all_tables(
        "Q1",
        q1
    )

    plot_q1(q1)

    q2 = solve_q2(
        positions
    )

    save_all_tables(
        "Q2",
        q2
    )
    print("问题2数据输出完成", flush=True)

    write_result_excel(
        RESULT2_TEMPLATE,
        OUT / "result2_final.xlsx",
        q2
    )
    print("问题2 Excel输出完成", flush=True)

    plot_layout(
        q2,
        "Q2_镜场布局.png"
    )
    print("问题2图像输出完成", flush=True)

    q3 = solve_q3(
        positions
    )

    save_all_tables(
        "Q3",
        q3
    )

    write_result_excel(
        RESULT3_TEMPLATE,
        OUT / "result3_final.xlsx",
        q3
    )

    plot_layout(
        q3,
        "Q3_镜场布局.png"
    )

    plot_q2_q3(
        q2,
        q3
    )

    print("\nMonte-Carlo收敛检验")

    mc2 = mc_convergence(q2)
    mc3 = mc_convergence(q3)

    mc2.to_csv(
        OUT / "Q2_MC收敛检验.csv",
        index=False,
        encoding="utf-8-sig"
    )

    mc3.to_csv(
        OUT / "Q3_MC收敛检验.csv",
        index=False,
        encoding="utf-8-sig"
    )

    print(f"Q2最大相对变化率 {mc2['相对变化率'].max():.4f}")
    print(f"Q3最大相对变化率 {mc3['相对变化率'].max():.4f}")


    print("\n网格收敛检验")

    grid2 = grid_convergence(q2)
    grid3 = grid_convergence(q3)

    grid2.to_csv(
        OUT / "Q2_网格收敛检验.csv",
        index=False,
        encoding="utf-8-sig"
    )

    grid3.to_csv(
        OUT / "Q3_网格收敛检验.csv",
        index=False,
        encoding="utf-8-sig"
    )

    summary = pd.DataFrame([
        {
            "问题": "问题1",
            "吸收塔x(m)": q1.tower_xy[0],
            "吸收塔y(m)": q1.tower_xy[1],
            "镜面数": q1.number,
            "总面积(m2)": q1.area,
            "年平均输出(MW)": q1.power,
            "单位面积指标(kW/m2)": q1.q
        },
        {
            "问题": "问题2",
            "吸收塔x(m)": q2.tower_xy[0],
            "吸收塔y(m)": q2.tower_xy[1],
            "镜面数": q2.number,
            "总面积(m2)": q2.area,
            "年平均输出(MW)": q2.power,
            "单位面积指标(kW/m2)": q2.q
        },
        {
            "问题": "问题3",
            "吸收塔x(m)": q3.tower_xy[0],
            "吸收塔y(m)": q3.tower_xy[1],
            "镜面数": q3.number,
            "总面积(m2)": q3.area,
            "年平均输出(MW)": q3.power,
            "单位面积指标(kW/m2)": q3.q
        }
    ])

    summary.to_csv(
        OUT / "最终结果汇总.csv",
        index=False,
        encoding="utf-8-sig"
    )

    print("\n运行结束")
    print("全部计算完成")
    print("A题 定日镜场优化设计")
    print(summary.to_string(index=False))
    print(f"\n结果目录：{OUT}")


if __name__ == "__main__":
    main()
