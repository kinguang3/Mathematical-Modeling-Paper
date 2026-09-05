from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy

from config import TABLE_DIR, OUTPUT_DIR
# 导入python库，config文件

# 写入excel的表单
def _write_sheet(ws, df):
    # 最上方添加列名
    ws.append(list(df.columns))
    # 取消索引，不命名元组
    for row in df.itertuples(index=False, name=None): 
        # 将dataframe数据写入excel
        ws.append(list(row))

    # 冻结表头，让整个表的第一行一直保持显示
    ws.freeze_panes="A2"
    #给数据添加自动删选功能
    ws.auto_filter.ref=ws.dimensions

    # 给第一行加粗
    for cell in ws[1]: 
        cell.font=copy(cell.font)
        cell.font=cell.font.copy(bold=True)

    # 自动调节列宽
    for col in range(1, ws.max_column + 1):
        width = min(28,max(10,max(
                (
                    len(str(ws.cell(r, col).value))
                    for r in range(1, min(ws.max_row, 100) + 1)
                ),
                default=10
            ) + 2
        )
    )
    ws.column_dimensions[get_column_letter(col)].width = width


def save_tables(tables):
    for name,df in tables.items():
        if isinstance(df,pd.DataFrame): df.to_csv(TABLE_DIR/f"{name}.csv",index=True,encoding="utf-8-sig")
    wb=Workbook(); wb.remove(wb.active)
    for name,df in tables.items():
        if not isinstance(df,pd.DataFrame): continue
        ws=wb.create_sheet(name[:31]); _write_sheet(ws,df.reset_index() if df.index.name or name=="table5" else df)
    wb.save(OUTPUT_DIR/"C题_全部计算表格.xlsx")


def save_run_summary(summary):
    (OUTPUT_DIR/"运行说明.txt").write_text(summary,encoding="utf-8")
