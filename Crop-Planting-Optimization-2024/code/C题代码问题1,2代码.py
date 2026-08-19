import os, copy
import numpy as np
import pandas as pd
from scipy.optimize import milp, LinearConstraint, Bounds
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
ATT1 = os.path.join(BASE, "附件1.xlsx")
ATT2 = os.path.join(BASE, "附件2.xlsx")
T11  = os.path.join(BASE, "result1_1.xlsx")
T12  = os.path.join(BASE, "result1_2.xlsx")
T2   = os.path.join(BASE, "result2.xlsx")

# 数据读取
land_df = pd.read_excel(ATT1, sheet_name=0).iloc[:54].copy()
land_df.columns = ["name","type","area","desc"]
land_df = land_df.dropna(subset=["name"])
land_df["type"] = land_df["type"].astype(str).str.strip()

crop_df = pd.read_excel(ATT1, sheet_name=1).iloc[:41].copy()
crop_df.columns = ["id","name","type","land","notes"]

stats_df = pd.read_excel(ATT2, sheet_name=1).iloc[:110].copy()
stats_df = stats_df.dropna(subset=["作物编号"])
stats_df.columns = ["seq","id","crop","land","season","yield","cost","price"]
stats_df = stats_df[pd.to_numeric(stats_df["id"], errors="coerce").notna()].copy()
stats_df["id"] = stats_df["id"].astype(int)
stats_df["land"] = stats_df["land"].astype(str).str.strip()
stats_df["season"] = stats_df["season"].astype(str).str.strip()
stats_df["price_mid"] = stats_df["price"].apply(
    lambda s: sum(map(float, str(s).split("-"))) / 2
)
statmap = {
    (int(r["id"]), r["land"], r["season"]):
    (float(r["yield"]), float(r["cost"]), float(r["price_mid"]))
    for _, r in stats_df.iterrows()
}

p23 = pd.read_excel(ATT2, sheet_name=0).iloc[:87].copy()
p23.columns = ["plot","id","crop","type","area","season"]
p23["plot"] = p23["plot"].ffill()
landmap = dict(zip(land_df["name"], land_df["type"]))

plots = [(r["name"], r["type"], float(r["area"])) for _,r in land_df.iterrows()]
crops = list(range(1,42))
bean = {1,2,3,4,5,17,18,19}
veg = set(range(17,35))
rad = {35,36,37}
mush = set(range(38,42))
seasons_by_type = {
    "平旱地":["单季"], "梯田":["单季"], "山坡地":["单季"],
    "水浇地":["单季","第一季","第二季"],
    "普通大棚":["第一季","第二季"],
    "智慧大棚":["第一季","第二季"]
}

demand = {i:0.0 for i in crops}
prev = {}
for _,r in p23.iterrows():
    if pd.isna(r["id"]): continue
    i = int(r["id"])
    p = str(r["plot"])
    s = str(r["season"]).strip()
    lt = landmap[p]
    if (i,lt,s) in statmap:
        demand[i] += float(r["area"]) * statmap[(i,lt,s)][0]
    prev.setdefault(p,set()).add(i)

allowed = []
for p,lt,a in plots:
    for s in seasons_by_type[lt]:
        for i in crops:
            key = (i,lt,s)
            if lt == "智慧大棚" and s == "第一季":
                key = (i,"普通大棚","第一季")
            if key in statmap:
                allowed.append((p,lt,a,s,i,key))

def solve_year(prev_sets, demand_y, scenario=2, cost_mult=1.0,
               price_mult=None, yield_mult=1.0, force_bean=None):
    nX = len(allowed)
    nY = 41
    nZ = 41 if scenario == 2 else 0
    nD = sum(lt=="水浇地" for _,lt,_ in plots)
    nB = len(plots)

    offY = nX
    offZ = offY+nY
    offD = offZ+nZ
    offB = offD+nD
    n = offB+nB

    c = np.zeros(n)
    integ = np.zeros(n)
    lb = np.zeros(n)
    ub = np.full(n, np.inf)

    water_plots = [p for p,lt,a in plots if lt=="水浇地"]

    for k,(p,lt,a,s,i,key) in enumerate(allowed):
        cost = statmap[key][1]
        c[k] = cost * cost_mult
        if i in prev_sets.get(p,set()):
            ub[k] = 0

    for i in crops:
        p = float(stats_df.loc[stats_df["id"]==i,"price_mid"].iloc[0])
        pm = 1.0 if price_mult is None else price_mult(i)
        c[offY+i-1] = -p*pm
        ub[offY+i-1] = demand_y[i]
        if scenario == 2:
            c[offZ+i-1] = -0.5*p*pm

    integ[offD:offD+nD] = 1
    ub[offD:offD+nD] = 1
    integ[offB:] = 1
    ub[offB:] = 1

    if force_bean:
        for bi,(p,lt,a) in enumerate(plots):
            if force_bean.get(p,False):
                lb[offB+bi] = 1

    cons, lo, hi = [], [], []

    # 面积约束
    for p,lt,a in plots:
        for s in seasons_by_type[lt]:
            row = np.zeros(n)
            for k,(pp,ltt,aa,ss,i,key) in enumerate(allowed):
                if pp==p and ss==s: row[k]=1
            cons.append(row); lo.append(-np.inf); hi.append(a)

    # 产量与销售量
    for i in crops:
        row = np.zeros(n)
        row[offY+i-1] = 1
        for k,(p,lt,a,s,ii,key) in enumerate(allowed):
            if ii==i:
                row[k] -= statmap[key][0]*yield_mult
        cons.append(row); lo.append(-np.inf); hi.append(0)

        if scenario == 2:
            row = np.zeros(n)
            row[offY+i-1] = 1
            row[offZ+i-1] = 1
            for k,(p,lt,a,s,ii,key) in enumerate(allowed):
                if ii==i:
                    row[k] -= statmap[key][0]*yield_mult
            cons.append(row); lo.append(0); hi.append(0)

    # 水浇地：水稻 vs 两季蔬菜
    for wi,p in enumerate(water_plots):
        a = next(a for pp,lt,a in plots if pp==p)
        d = offD+wi

        row=np.zeros(n)
        for k,(pp,ltt,aa,s,i,key) in enumerate(allowed):
            if pp==p and s=="单季" and i==16: row[k]=1
        row[d]=-a
        cons.append(row); lo.append(-np.inf); hi.append(0)

        row=np.zeros(n)
        for k,(pp,ltt,aa,s,i,key) in enumerate(allowed):
            if pp==p and s=="第一季" and i in veg: row[k]=1
        row[d]=a
        cons.append(row); lo.append(-np.inf); hi.append(a)

        row=np.zeros(n)
        for k,(pp,ltt,aa,s,i,key) in enumerate(allowed):
            if pp==p and s=="第二季" and i in rad: row[k]=1
        row[d]=a
        cons.append(row); lo.append(-np.inf); hi.append(a)

    # 豆类三年轮作：b=1表示该地块本年种过豆类
    for bi,(p,lt,a) in enumerate(plots):
        b = offB+bi
        row=np.zeros(n); row[b]=-a
        for k,(pp,ltt,aa,s,i,key) in enumerate(allowed):
            if pp==p and i in bean: row[k]=1
        cons.append(row); lo.append(-np.inf); hi.append(0)

        row=np.zeros(n); row[b]=0.01
        for k,(pp,ltt,aa,s,i,key) in enumerate(allowed):
            if pp==p and i in bean: row[k]-=1
        cons.append(row); lo.append(-np.inf); hi.append(0)

    res = milp(
        c, integrality=integ, bounds=Bounds(lb,ub),
        constraints=LinearConstraint(
            np.array(cons), np.array(lo), np.array(hi)
        ),
        options={"time_limit":120, "mip_rel_gap":1e-5}
    )
    if not res.success:
        raise RuntimeError(res.message)
    return res

def run_scenario(scenario=2, q2=False):
    prev_sets = copy.deepcopy(prev)
    bean_hist = {
        p:[int(any(i in bean for i in prev_sets.get(p,set())))]
        for p,_,_ in plots
    }
    solutions, profits = {}, {}

    for year in range(2024,2031):
        if q2:
            demand_y = {
                i: demand[i]*(1.075**(year-2023) if i in {6,7} else 1.0)
                for i in crops
            }
            def pm(i):
                typ = str(crop_df.loc[crop_df["id"]==i,"type"].iloc[0])
                if "蔬菜" in typ: return 1.05**(year-2023)
                if i==41: return 0.95**(year-2023)
                if i in mush: return 0.97**(year-2023)
                return 1.0
            cm = 1.05**(year-2023)
        else:
            demand_y = demand.copy()
            pm = lambda i: 1.0
            cm = 1.0

        force = {
            p:(len(bean_hist[p])>=2 and sum(bean_hist[p][-2:])==0)
            for p,_,_ in plots
        }

        res = solve_year(
            prev_sets, demand_y, scenario, cm, pm, force_bean=force
        )
        x = res.x[:len(allowed)]
        sol = [
            (a[0],a[3],a[4],x[k])
            for k,a in enumerate(allowed) if x[k] > 1e-6
        ]
        solutions[year] = sol
        profits[year] = -res.fun

        prev_sets = {p:set() for p,_,_ in plots}
        for p,s,i,v in sol:
            prev_sets[p].add(i)
        for p,_,_ in plots:
            bean_hist[p].append(int(any(i in bean for i in prev_sets[p])))

    return solutions, profits

def write_solution(template, out_file, solution):
    wb = load_workbook(template)
    crop_cols = {
        str(wb["2024"].cell(1,c).value).strip():c
        for c in range(3,44) if wb["2024"].cell(1,c).value is not None
    }
    crop_name = {int(r["id"]):str(r["name"]).strip()
                 for _,r in crop_df.iterrows()}

    for year in range(2024,2031):
        ws = wb[str(year)]
        row_first = {str(ws.cell(r,2).value).strip():r
                     for r in range(2,56) if ws.cell(r,2).value is not None}
        row_second = {str(ws.cell(r,2).value).strip():r
                      for r in range(56,84) if ws.cell(r,2).value is not None}
        for plot,season,cid,area in solution[year]:
            name = crop_name[int(cid)]
            col = crop_cols.get(name)
            if col is None:
                m=[k for k in crop_cols if k.strip()==name]
                if not m: continue
                col=crop_cols[m[0]]
            row = row_second.get(plot) if season=="第二季" else row_first.get(plot)
            if row is not None:
                ws.cell(row,col).value = round(float(area),4)
    wb.save(out_file)

if __name__ == "__main__":
    s11,p11 = run_scenario(1,False)
    s12,p12 = run_scenario(2,False)
    s2,p2   = run_scenario(2,True)

    write_solution(T11, os.path.join(BASE,"result1_1_python运行结果.xlsx"), s11)
    write_solution(T12, os.path.join(BASE,"result1_2_python运行结果.xlsx"), s12)
    write_solution(T2,  os.path.join(BASE,"result2_python运行结果.xlsx"), s2)

    print("问题1(1)总利润:",sum(p11.values()))
    print("问题1(2)总利润:",sum(p12.values()))
    print("问题2总利润:",sum(p2.values()))
