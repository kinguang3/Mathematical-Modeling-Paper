# -*- coding: utf-8 -*-
"""
C题问题3：1000次蒙特卡洛 + CVaR完整实现
需要：附件1.xlsx、附件2(1).xlsx、result2_python运行结果.xlsx
报告参数：N=1000；Corr(销量,价格)=-0.6；Corr(成本,价格)=0.3；
豆类→非豆类产量加成5%；lambda=0.2；alpha=0.95。
"""
# 该文件为本次实际运行的完整脚本版本，算法步骤与生成结果一致。
# 为保持代码文件可独立复现，请将本文件与3个输入Excel放在同一目录后运行。
import os,re,numpy as np,pandas as pd
from openpyxl import load_workbook,Workbook
BASE=os.path.dirname(os.path.abspath(__file__))
ATT1=os.path.join(BASE,"附件1.xlsx"); ATT2=os.path.join(BASE,"附件2.xlsx"); R2=os.path.join(BASE,"result2_python运行结果.xlsx")
land=pd.read_excel(ATT1,sheet_name=0).iloc[:54].copy(); land.columns=["name","type","area","desc"]; land=land.dropna(subset=["name"]); land["type"]=land["type"].astype(str).str.strip()
crop=pd.read_excel(ATT1,sheet_name=1).iloc[:41].copy(); crop.columns=["id","name","type","land","notes"]; crop["id"]=pd.to_numeric(crop["id"],errors="coerce"); crop=crop.dropna(subset=["id"]); crop["id"]=crop["id"].astype(int)
raw=pd.read_excel(ATT2,sheet_name=1); raw=raw.rename(columns={raw.columns[1]:"id"}); raw["id_num"]=pd.to_numeric(raw["id"],errors="coerce")
st=raw[raw["id_num"].notna()].iloc[:110,:8].copy(); st.columns=["seq","id","crop","land","season","yield","cost","price"]; st["id"]=pd.to_numeric(st["id"],errors="coerce").astype(int); st["land"]=st["land"].astype(str).str.strip(); st["season"]=st["season"].astype(str).str.strip(); st["yield"]=pd.to_numeric(st["yield"],errors="coerce"); st["cost"]=pd.to_numeric(st["cost"],errors="coerce")
def pm(x):
    a=[float(v) for v in re.findall(r"[\d.]+",str(x))]; return sum(a)/len(a)
st["price_mid"]=st["price"].map(pm); st=st.dropna(subset=["yield","cost","price_mid"])
stat={(int(r.id),r.land,r.season):(float(r["yield"]),float(r["cost"]),float(r["price_mid"])) for _,r in st.iterrows()}; landmap=dict(zip(land["name"],land["type"]))
p23=pd.read_excel(ATT2,sheet_name=0).iloc[:87].copy(); p23.columns=["plot","id","crop","type","area","season"]; p23["plot"]=p23["plot"].ffill(); p23["id_num"]=pd.to_numeric(p23["id"],errors="coerce"); p23["area_num"]=pd.to_numeric(p23["area"],errors="coerce")
demand={i:0. for i in range(1,42)}
# ==========仅此处修改 r.plot → r["plot"]，其余全部原样保留==========
for _,r in p23[p23["id_num"].notna()].iterrows():
    i=int(r.id_num); p=str(r["plot"]).strip(); s=str(r.season).strip(); lt=landmap[p]
# =================================================================
    if (i,lt,s) in stat: demand[i]+=float(r.area_num)*stat[(i,lt,s)][0]
wb=load_workbook(R2,data_only=True); name_to_id={str(r["name"]).strip():int(r["id"]) for _,r in crop.iterrows()}; solutions={}
for year in range(2024,2031):
    ws=wb[str(year)]; cols={}
    for c in range(1,ws.max_column+1):
        v=ws.cell(1,c).value
        if v is not None and str(v).strip() in name_to_id: cols[str(v).strip()]=c
    if not cols:
        for c in range(3,min(43,ws.max_column+1)):
            if c-3<len(crop): cols[str(crop.iloc[c-3]["name"]).strip()]=c
    arr=[]
    for rr in range(2,ws.max_row+1):
        p=ws.cell(rr,2).value
        if p is None: continue
        p=str(p).strip(); lt=landmap.get(p)
        if lt is None: continue
        second=rr>=56
        for name,col in cols.items():
            try:a=float(ws.cell(rr,col).value)
            except:continue
            if a<=1e-8:continue
            i=name_to_id[name]; s="第二季" if second else ("单季" if lt=="水浇地" and i==16 else "第一季")
            if (i,lt,s) in stat:arr.append((p,s,i,a))
    solutions[year]=arr
bean={1,2,3,4,5,17,18,19}; veg=set(range(17,35)); mush={38,39,40,41}
N=1000; SEED=20240813; LAMBDA=.2; ALPHA=.95
rng=np.random.default_rng(SEED); R=np.array([[1,-.6,0],[-.6,1,.3],[0,.3,1.]]); Z=rng.normal(size=(N,3))@np.linalg.cholesky(R).T
def scen(year,z):
    k=year-2023;sales={};price={};cost={};ym={}
    for i in range(1,42):
        sales[i]=1+np.clip(.075+.025*z[0],.05,.10) if i in (6,7) else 1+np.clip(.05*z[0],-.05,.05)
        ym[i]=1+np.clip(.10*z[2],-.10,.10)
        base=1.05**k if i in veg else (.95**k if i==41 else (.97**k if i in mush else 1.))
        price[i]=max(.01,base*(1+.10*z[1])); cost[i]=1.05**k*(1+.05*z[1])
    return sales,price,cost,ym
def profit(sol,year,s):
    sales,price,cost,ym=s;q={i:0. for i in range(1,42)};tc=0.
    for p,se,i,a in sol:
        key=(i,landmap[p],se)
        if key not in stat:continue
        y,c,p0=stat[key];q[i]+=a*y*ym[i]*(1.05 if i not in bean else 1.);tc+=a*c*cost[i]
    rev=0.
    for i,v in q.items():
        D=demand[i]*sales[i];P=float(st.loc[st.id==i,"price_mid"].iloc[0])*price[i];rev+=P*min(v,D)+.5*P*max(v-D,0)
    return rev-tc
annual=np.zeros((N,7))
for k in range(N):
    for j,y in enumerate(range(2024,2031)):annual[k,j]=profit(solutions[y],y,scen(y,Z[k]))
total=annual.sum(axis=1);p5=float(np.quantile(total,.05));cvar=float(total[total<=p5].mean());exp=float(total.mean());risk=exp-LAMBDA*(exp-cvar)
out=os.path.join(BASE,"问题3_完整蒙特卡洛_CVaR结果.xlsx");w=Workbook();ws=w.active;ws.title="汇总"
for r in [["指标","数值"],["模拟次数",N],["Corr(销量,价格)",-.6],["Corr(成本,价格)",.3],["豆类→非豆类产量加成",.05],["lambda",LAMBDA],["alpha",ALPHA],["期望利润(万元)",exp/1e4],["P5利润(万元)",p5/1e4],["CVaR_0.95(万元)",cvar/1e4],["风险调整目标(万元)",risk/1e4]]:ws.append(r)
ws2=w.create_sheet("1000场景");ws2.append(["场景"]+[str(y) for y in range(2024,2031)]+["合计","尾部"])
for k in range(N):ws2.append([k+1]+list(annual[k])+[total[k],int(total[k]<=p5)])
w.save(out);print(out,exp/1e4,cvar/1e4)